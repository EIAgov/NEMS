# -*- coding: utf-8 -*-
"""
Created POC code on Dec 18, 2023
Developed Prototype on Jan 8, 2024
Developed Alpha version on Feb 6, 2024, 1/3 table programs
Developed Beta version on Mar 6, 2024, 1/3 table programs
Updated on Apr 3, 2024: 1/3 table programs, 
Updated on May 6, 2024: RAN, longthin, debugging
Updated on Jun 6, 2024: aeo table format/style  
Updated on Jun 20, 2024: output control
Updated on Jul 1, 2024: fixing data issues 

@author: SZO
@author2: JMW
"""

import sys
import os
import time
import calendar

import csv
import tomllib
import traceback
from configparser import ConfigParser
from sys import argv

import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

from RW_preprocessor import (
    parse_parameter_file,
    get_table_format_csv,
    get_table_format,
    normalize_format,
    insert_citation,
    sum_CarbCap_vars,
)
from RW_preprocessor_sme import preprocessor_sme
from RW_postprocessor_base import postprocessor_base
from RW_make_ran import make_ran

# Specify the path to the folder containing base table programs
sys.path.append(os.path.join(os.getcwd(), "RW tables"))
table_programs_path = os.getcwd() + "\\reporter\\RW tables"
sys.path.append(os.getcwd() + "\\reporter\\input")
sys.path.append(os.getcwd() + "\\reporter\\output")
sys.path.append(os.getcwd() + "\\reporter\\RW tables")
sys.path.append(os.getcwd() + "\\PyFiler")
sys.path.append(os.getcwd() + "\\RW Tables")
sys.path.append(os.getcwd() + "\\input")
sys.path.append(os.getcwd() + "\\output")
# Add the path to sys.path
sys.path.append(table_programs_path)

# pylint: disable=import-error
from RW_fill_table_base_001 import fill_table_base_001
from RW_fill_table_base_002 import fill_table_base_002
from RW_fill_table_base_003 import fill_table_base_003
from RW_fill_table_base_004 import fill_table_base_004
from RW_fill_table_base_005 import fill_table_base_005
from RW_fill_table_base_006 import fill_table_base_006
from RW_fill_table_base_007 import fill_table_base_007
from RW_fill_table_base_008 import fill_table_base_008
from RW_fill_table_base_009 import fill_table_base_009
from RW_fill_table_base_010 import fill_table_base_010
from RW_fill_table_base_011 import fill_table_base_011
from RW_fill_table_base_012 import fill_table_base_012
from RW_fill_table_base_013 import fill_table_base_013
from RW_fill_table_base_014 import fill_table_base_014
from RW_fill_table_base_015 import fill_table_base_015
from RW_fill_table_base_016 import fill_table_base_016
from RW_fill_table_base_017 import fill_table_base_017
from RW_fill_table_base_018 import fill_table_base_018
from RW_fill_table_base_019 import fill_table_base_019
from RW_fill_table_base_020 import fill_table_base_020
from RW_fill_table_base_021 import fill_table_base_021
from RW_fill_table_base_022 import fill_table_base_022
from RW_fill_table_base_023 import fill_table_base_023
from RW_fill_table_base_024 import fill_table_base_024
from RW_fill_table_base_025 import fill_table_base_025
from RW_fill_table_base_026 import fill_table_base_026
from RW_fill_table_base_027 import fill_table_base_027
from RW_fill_table_base_028 import fill_table_base_028
from RW_fill_table_base_029 import fill_table_base_029
from RW_fill_table_base_030 import fill_table_base_030
from RW_fill_table_base_031 import fill_table_base_031
from RW_fill_table_base_032 import fill_table_base_032
from RW_fill_table_base_033 import fill_table_base_033
from RW_fill_table_base_034 import fill_table_base_034
from RW_fill_table_base_035 import fill_table_base_035
from RW_fill_table_base_036 import fill_table_base_036
from RW_fill_table_base_037 import fill_table_base_037
from RW_fill_table_base_038 import fill_table_base_038
from RW_fill_table_base_039 import fill_table_base_039
from RW_fill_table_base_040 import fill_table_base_040
from RW_fill_table_base_041 import fill_table_base_041
from RW_fill_table_base_042 import fill_table_base_042
from RW_fill_table_base_043 import fill_table_base_043
from RW_fill_table_base_044 import fill_table_base_044
from RW_fill_table_base_045 import fill_table_base_045
from RW_fill_table_base_046 import fill_table_base_046
from RW_fill_table_base_047 import fill_table_base_047
from RW_fill_table_base_048 import fill_table_base_048
from RW_fill_table_base_049 import fill_table_base_049
from RW_fill_table_base_050 import fill_table_base_050
from RW_fill_table_base_051 import fill_table_base_051
from RW_fill_table_base_052 import fill_table_base_052
from RW_fill_table_base_053 import fill_table_base_053
from RW_fill_table_base_054 import fill_table_base_054
from RW_fill_table_base_055 import fill_table_base_055
from RW_fill_table_base_056 import fill_table_base_056
from RW_fill_table_base_057 import fill_table_base_057
from RW_fill_table_base_058 import fill_table_base_058
from RW_fill_table_base_059 import fill_table_base_059
from RW_fill_table_base_060 import fill_table_base_060
from RW_fill_table_base_061 import fill_table_base_061
from RW_fill_table_base_062 import fill_table_base_062

#from RW_fill_table_base_063 import fill_table_base_063 #BLANK TABLE NOW
#from RW_fill_table_base_064 import fill_table_base_064 #BLANK TABLE NOW
from RW_fill_table_base_065 import fill_table_base_065
from RW_fill_table_base_066 import fill_table_base_066
from RW_fill_table_base_067 import fill_table_base_067
from RW_fill_table_base_068 import fill_table_base_068
from RW_fill_table_base_069 import fill_table_base_069
from RW_fill_table_base_070 import fill_table_base_070
from RW_fill_table_base_071 import fill_table_base_071
from RW_fill_table_base_072 import fill_table_base_072
from RW_fill_table_base_073 import fill_table_base_073
from RW_fill_table_base_074 import fill_table_base_074
from RW_fill_table_base_075 import fill_table_base_075
from RW_fill_table_base_076 import fill_table_base_076
from RW_fill_table_base_077 import fill_table_base_077
from RW_fill_table_base_078 import fill_table_base_078
from RW_fill_table_base_079 import fill_table_base_079
from RW_fill_table_base_080 import fill_table_base_080
from RW_fill_table_base_081 import fill_table_base_081

# from RW_fill_table_base_082  import fill_table_base_082  # blank
from RW_fill_table_base_083 import fill_table_base_083

# from RW_fill_table_base_084  import fill_table_base_084  # blank
# from RW_fill_table_base_085  import fill_table_base_085  # blank
# from RW_fill_table_base_086  import fill_table_base_086  # blank
# from RW_fill_table_base_087  import fill_table_base_087  # blank
# from RW_fill_table_base_088  import fill_table_base_088  # blank
# from RW_fill_table_base_089  import fill_table_base_089  # blank
from RW_fill_table_base_090 import fill_table_base_090
from RW_fill_table_base_091 import fill_table_base_091

# from RW_fill_table_base_092  import fill_table_base_092  # blank
from RW_fill_table_base_093 import fill_table_base_093
from RW_fill_table_base_094 import fill_table_base_094
from RW_fill_table_base_095 import fill_table_base_095
from RW_fill_table_base_096 import fill_table_base_096
from RW_fill_table_base_097 import fill_table_base_097
from RW_fill_table_base_098 import fill_table_base_098
from RW_fill_table_base_099 import fill_table_base_099
from RW_fill_table_base_100 import fill_table_base_100
from RW_fill_table_base_101 import fill_table_base_101
from RW_fill_table_base_102 import fill_table_base_102
from RW_fill_table_base_103 import fill_table_base_103

# from RW_fill_table_base_104  import fill_table_base_104  # blank
from RW_fill_table_base_105 import fill_table_base_105
from RW_fill_table_base_106 import fill_table_base_106
from RW_fill_table_base_107 import fill_table_base_107
from RW_fill_table_base_108 import fill_table_base_108
#from RW_fill_table_base_109 import fill_table_base_109 #BLANK TABLE NOW
from RW_fill_table_base_110 import fill_table_base_110
from RW_fill_table_base_111 import fill_table_base_111
from RW_fill_table_base_112 import fill_table_base_112
from RW_fill_table_base_113 import fill_table_base_113
from RW_fill_table_base_114 import fill_table_base_114
from RW_fill_table_base_115 import fill_table_base_115
from RW_fill_table_base_116 import fill_table_base_116
from RW_fill_table_base_117 import fill_table_base_117
from RW_fill_table_base_118 import fill_table_base_118
from RW_fill_table_base_119 import fill_table_base_119

# from RW_fill_table_base_120  import fill_table_base_120  # blank
# from RW_fill_table_base_121  import fill_table_base_121  # blank
from RW_fill_table_base_122 import fill_table_base_122
from RW_fill_table_base_123 import fill_table_base_123
from RW_fill_table_base_124 import fill_table_base_124
from RW_fill_table_base_125 import fill_table_base_125
from RW_fill_table_base_126 import fill_table_base_126
from RW_fill_table_base_127 import fill_table_base_127
from RW_fill_table_base_128 import fill_table_base_128
from RW_fill_table_base_129 import fill_table_base_129
from RW_fill_table_base_130 import fill_table_base_130
from RW_fill_table_base_131 import fill_table_base_131
from RW_fill_table_base_132 import fill_table_base_132
from RW_fill_table_base_133 import fill_table_base_133
from RW_fill_table_base_134 import fill_table_base_134
from RW_fill_table_base_135 import fill_table_base_135
from RW_fill_table_base_136 import fill_table_base_136
from RW_fill_table_base_137 import fill_table_base_137
from RW_fill_table_base_138 import fill_table_base_138
from RW_fill_table_base_139 import fill_table_base_139
from RW_fill_table_base_140 import fill_table_base_140
from RW_fill_table_base_141 import fill_table_base_141
from RW_fill_table_base_142 import fill_table_base_142
from RW_fill_table_base_143 import fill_table_base_143
from RW_fill_table_base_144 import fill_table_base_144
from RW_fill_table_base_145 import fill_table_base_145
from RW_fill_table_base_146 import fill_table_base_146
from RW_fill_table_base_147 import fill_table_base_147

from RW_fill_table_base_148 import fill_table_base_148
from RW_fill_table_base_149 import fill_table_base_149
from RW_fill_table_base_150 import fill_table_base_150

from RW_debug import main_logger, log_execution

# Get settings from config.ini
config = ConfigParser()
# print(os.getcwd())
if os.getcwd()[-8:] != "reporter":
    configpath = os.getcwd() + "\\reporter\\config.ini"
else:
    configpath = os.getcwd() + "\\config.ini"
config.read(configpath)
config.sections()


@log_execution(main_logger, message="init")
def init(user):
    """Initializes and returns a dictionary called table_spec. This dictionary
    serves as metadata specifying various inputs and info related to tables,
    including interface (connection) to NEMS database restart.npz

    Parameters
    ----------
    user : user
        User object containing user-specific information - see user class

    Returns
    -------
    table_spec : dict =====================================================
        Serves as metadata specifying all table related info, excluding data

        Settings for table print related controls -------------------------
        'settings'              # Table print controls (ftab.dat)

        Table number mapping between aeo published tables and layin tables
        'table_mapping'         # e.g, aeo table 18 <=> layin table 17

        General NEMS info--------------------------------------------------
        'StudyID'		        # e.g, 'AEO2023'
        'LongTitle'		        # e.g, 'Annual Energy Outlook 2023'
        'case_descriptions'	    # e.g, 'Reference Case'
        'ScenLabel'		        # e.g, 'Reference'
        'Scenario		        # e.g, 'ref2024'
        'DateKey'		        # e.g, 'd101323b'
        'ScenDate'		        # e.g, 'ref2023.d020623a'
        'case_descriptions'	    # e.g, 'Reference Case'

        Projection years---------------------------------------------------
        'first_year'		    # e.g, 1990
        'last_year'		        # e.g, 2050
        'years			        # e.g, [1990, 1991, ..., 2050]
        'years_str		        # e.g, ['1990', '1991', ..., '2050']

        Print years -------------------------------------------------------
        'first_print_year'	    # e.g, 2020
        'last_print_year'	    # e.g, 2050
        'print_years'		    # e.g, [2020, 2021, ..., 2050]

        Print regions -----------------------------------------------------
        (Note: Some Supplemental files have regional tables)
        'print_regions'         # 1: print regions, 0, don't print

        Input files--------------------------------------------------------
        'table_ids_full'	    # list of all tables available
        'table_ids'	            # list of tables to print
        'table_var_def_path'	# input file of 'table_var_def.toml'
        'df_format'	            # df of normalized layin
        'table_input'	        # dict: {TableID: {{Regionality}, Program}}
        'table_input_df'	    # the above dict in df (easy to use)
        'dat'	                # connection/interface to NEMS restart.npz
        'coefficients_path'     # input file of 'RW_coefficients.csv'
        'layin_ver'	            # switch to use 'layin.csv' or 'layin.xls'
        'citation_path'	        # input file of 'citation.txt'
        'tabreq_path'           # input file of 'tabreq.txt'

        Additional info required for RAN-----------------------------------
        'RSVe			        # 'AA01' for Scenario Header data
        'RSFile'		        # 'restart.npz'
        'RSScen'		        # '../../ref2024/d101323b/RESTART.unf'
        'all_row_file_path'     # all row csv file, raw data for make_ran

    dat : dataframe =====================================================
        represents an connection/interface to NEMS restart.npz

    """

    # Set input files: get from user
    restartnpz_path = os.path.join(os.getcwd(), user.restartnpz_path)
    table_var_def_path = os.path.join(os.getcwd(), user.table_var_def_path)
    layin_ver = user.layin_ver
    format_path = os.path.join(os.getcwd(), user.format_path)
    format_path_csv = os.path.join(os.getcwd(), user.format_path_csv)
    table_input_path = os.path.join(os.getcwd(), user.table_input_path)
    region_input_path = os.path.join(os.getcwd(), user.region_input_path)
    coefficients_path = os.path.join(os.getcwd(), user.coefficients_path)
    citation_path = os.path.join(os.getcwd(), user.citation_path)
    tabreq_path = os.path.join(os.getcwd(), user.tabreq_path)
    table_mapping_path = os.path.join(os.getcwd(), user.table_mapping_path)

    # Check table availablility to get a full list of all available tables
    table_num_full = list(set(range(1, 151)).difference(set(user.unavailable_tables)))
    table_ids_full = ["TN " + str(num).zfill(3) for num in table_num_full]

    # Get tables required
    table_ids_req = read_tabreq(tabreq_path)

    # Update table_ids_req to make sure a required table is available in tabel_ids_full
    table_ids_req = [i for i in table_ids_req if i in table_ids_full]

    main_logger.info("\n============ table_ids_full ============")
    main_logger.info(table_ids_full)
    main_logger.info("\n============ table_ids_req ============")
    main_logger.info(table_ids_req)

    # Read table print settings from config.ini
    settings = {}
    inputpath = {}
    casedescription = {}
    outputdir = {}
    for key, value in config.items("settings"):
        settings[key] = value
    for key, value in config.items("inputpath"):
        inputpath[key] = value
    for key, value in config.items("casedescriptions"):
        casedescription[key] = value
    for key, value in config.items("outputpath"):
        outputdir[key] = value

    # ''' Table print settings in config.ini (for users to control table output)
    # -----------------------------------------------------------------------------------------
    # [settings]
    #     print_aeotab1_20 = 1                flag 1: print, 0: don't print
    #     print_yearbyyear = 1                flag 1: print, 0: don't print
    #     print_supplemental_files = 1        flag 1: print, 0: don't print
    #     print_regional_tables = 1           flag 1: print, 0: don't print
    #     create_longthin_files = 1           flag 1: print, 0: don't print
    #     create_ran_file = 1                 flag 1: print, 0: don't print

    #     print_base_tables = 0               flag 1: print, 0: don't print
    #     print_tabreq_tables = 0             flag 1: print, 0: don't print
    #     print_tabreq_tables_stack = 0       flag 1: print, 0: don't print

    # -----------------------------------------------------------------------------------------
    #     max_col = 34                        MAXCOL        MAX COLS/PG
    #     i_col_width = 6                     ICOLWD    REPORT COLUMN WIDTH
    #     print_regions = 0                   IPREGD   PRINT REGIONS (1=YES)
    #     print_footnotes = 0                 IPRFTN  PRINT FOOTNOTES (1=YES)
    #     first_print_year = 2000             IFSTYR  FIRST PRINT YEAR
    #     last_print_year = 2050              ILSTYR  LAST PRINT YEAR
    #     growth_rate_option = 0              IGROWO,IGRWYR GROWTH RATE OPTION (1=YES)
    #     growth_rate_start = 2022            IGROWO,IGRWYR GROWTH RATE OPTION START YEAR
    #     print_every_5year = 1               ISKIPO,ISKIPY 5-YEAR OPTION (1=YES)
    #     print_5year_start = 2050            ISKIPO,ISKIPY 5-YEAR OPTION START YEAR
    #     year_index_real_dol_priceyr = 2022  FYRPRC YEAR INDEX FOR REAL $ PRICE YEAR
    #     year_cum_cap_addition_begin = 2022  Year cumulative capacity additions begin from
    #     year_cum_oilgas_prod_begin = 2022   Year cumulative oil and gas production begins
    #     carbon_or_carbondioxide = co2       Print carbon or carbon dioxide
    #     print_bonus_rows = 1                Print bonus rows as designated by DRForm in layout
    #     discount_rate_table116 = 0.08       Discount rate for table 116
    #     --------------------------------------------------------------------------------------
    # '''

    # Get projection years from user:
    # e.g., [1990, 2021, ..., 2050]
    years = list(range(user.first_year, user.last_year + 1))
    # e.g., ['1990', '1991', ..., '2050']
    years_str = [str(year) for year in years]

    # Get print year from config.ini:
    if SCEDES["SCEN"] != "ref2023":
        first_print_year = int(settings["first_print_year"])
        last_print_year = int(settings["last_print_year"])
        print_years_range = range(
            int(settings["first_print_year"]), int(settings["last_print_year"]) + 1
            )
    else:
        first_print_year = int(settings["first_print_year"])-1
        last_print_year = int(settings["last_print_year"])
        print_years_range = range(
            int(settings["first_print_year"])-1, int(settings["last_print_year"]) + 1
            )

    # e.g., [2020, 2021, ..., 2050]
    print_years = list(print_years_range)

    # Get table format from layin
    #   switch between layin.csv and layin.xls
    if user.layin_ver == "csv":
        df_format = get_table_format_csv(format_path_csv)
    else:
        df_format = get_table_format(format_path)

    # Normalize df_format to get three HD rows, and one RH row for all tables
    df_format = normalize_format(df_format, HD_num=3)

    # Insert citations
    df_format, MACYR = insert_citation(df_format, citation_path,user)
    df_format = df_format.reset_index(drop=True)

    # Updated on 4/29/2023 per EIA's request to store table input in csv files as:
    #   table_input.csv
    #   regions.txt
    table_input_df = pd.read_csv(table_input_path, index_col="TableID")
    table_input = table_input_df.T.to_dict()

    # Set regions
    df = pd.read_csv(region_input_path)

    # National tables
    df_r = df[df["Regionality"] == "NATNAM"]
    NATNAM = df_r.set_index("Region_ID")["Region_Name"].to_dict()

    # Census regions
    df_r = df[df["Regionality"] == "REGNAM"]
    REGNAM = df_r.set_index("Region_ID")["Region_Name"].to_dict()

    # EMM/NERC region
    df_r = df[df["Regionality"] == "NERCNAM"]
    NERCNAM = df_r.set_index("Region_ID")["Region_Name"].to_dict()

    # Region Spec
    # ''' National----------
    # NATNAM = {
    #     0: 'United States'
    #     }
    # '''

    # ''' Census Regions:---------
    # REGNAM = {
    #     1: 'New England',
    #     2: 'Middle Atlantic',
    #     3: 'East North Central',
    #     4: 'West North Central',
    #     5: 'South Atlantic',
    #     6: 'East South Central',
    #     7: 'West South Central',
    #     8: 'Mountain',
    #     9: 'Pacific',
    #     10: 'California',
    #     11: 'United States',
    #     }
    # '''

    # ''' EMM Regions: --------------------------------------------------------------
    # NERCNAM = {
    #     1: 'Texas Reliability Entity',
    #     2: 'Florida Reliability Coordinating Council',
    #     3: 'Midcontinent / West',
    #     4: 'Midcontinent / Central',
    #     5: 'Midcontinent / East',
    #     6: 'Midcontinent / South',
    #     7: 'Northeast Power Coordinating Council / New England',
    #     8: 'Northeast Power Coordinating Council / New York City and Long Island',
    #     9: 'Northeast Power Coordinating Council / Upstate New York',
    #     10: 'PJM / East',
    #     11: 'PJM / West',
    #     12: 'PJM / Commonwealth Edison',
    #     13: 'PJM / Dominion',
    #     14: 'SERC Reliability Corporation / East',
    #     15: 'SERC Reliability Corporation / Southeastern',
    #     16: 'SERC Reliability Corporation / Central',
    #     17: 'Southwest Power Pool / South',
    #     18: 'Southwest Power Pool / Central',
    #     19: 'Southwest Power Pool / North',
    #     20: 'Western Electricity Coordinating Council / Southwest',
    #     21: 'Western Electricity Coordinating Council / California North',
    #     22: 'Western Electricity Coordinating Council / California South',
    #     23: 'Western Electricity Coordinating Council / Northwest Power Pool Area',
    #     24: 'Western Electricity Coordinating Council / Rockies',
    #     25: 'Western Electricity Coordinating Council / Basin',
    #     26: 'Alaska',
    #     27: 'Hawaii',
    #     28: 'United States'
    #     }
    # '''
    # Populate actual region data to table_input
    for key, value in table_input.items():
        for inner_key, inner_value in value.items():
            if inner_value == "NATNAM":
                value[inner_key] = NATNAM
            elif inner_value == "REGNAM":
                value[inner_key] = REGNAM
            elif inner_value == "NERCNAM":
                value[inner_key] = NERCNAM

    # After the above operations, the dict table_input is structured as below:
    # ''' ---------------------------------------------------------------------
    # {'TN 001': {
    #             'Region' : {0: 'United States'},
    #             'Caption': 'Total Energy Supply, Disposition, and Price Summary',
    #             'Program': 'fill_table_base_001'
    #            },
    #  'TN 002': {
    #             'Region' : {1: 'New England', 2: 'Middle Atlantic',..., 11: United States},
    #             'Caption': 'Energy Consumption by Sector and Source',
    #             'Program': 'fill_table_base_002'
    #            },
    #  ......

    #  'TN 059': {
    #             'Region' : {1: 'Texas Reliability Entity', 2: 'Florida Reliability Coordinating Council',...,
    #                         28: 'United States'},
    #             'Caption': 'Electric Power Sector Generating Capacity and Generation by Plant Type and Technology',
    #             'Program': 'fill_table_base_059'
    #            },
    #  ......
    # }
    # ------------------------------------------------------------------------'''

    # Get all available tables (some tables may not be available: e.g., still under development)
    table_input = {
        key: table_input[key] for key in table_ids_full if key in table_input
    }

    # Get table mapping
    df = pd.read_excel(table_mapping_path)
    table_mapping = dict(zip(df["Layin Table Number"], df["Published Table Number"]))

    # Set connection / interface to NEMS database restart.npzs
    global dfd, dfd1

    if SCEDES["SCEN"] != "ref2023":
        dat = dfd = read_npz(user)
        dfdkeys = pd.DataFrame(dfd.keys())
        dfdkeys[["commonblock", "varname"]] = dfdkeys[0].str.split("/", expand=True)
        duplicate_varsnames = dfdkeys["varname"].duplicated(keep=False)
        dfdkeys[0] = dfdkeys[0].mask(duplicate_varsnames, dfdkeys[0])
        dfdkeys[0] = dfdkeys[0].where(duplicate_varsnames, dfdkeys["varname"])
        dfdkeys = dfdkeys.drop(columns=["commonblock", "varname"])
        dfdkeys = dfdkeys.set_index([0], drop=False)
        dfdhold = dfd.copy()

        for i in list(dfd):
            try:
                dfdkeys.loc[i]
            except KeyError:
                parts = i.split("/")
                dfdhold[parts[-1].upper()] = dfdhold.pop(i)

        dfd = dfdhold.copy()

        # Specifically used for AEO2025 for summing Carbon Capture Variables. Remove in AEO2026
        dfd = sum_CarbCap_vars(dfd)

        # Set connection / interface to NEMS database restart.npzs
    else:
        # In this piece of code, specifically for AEO2025, with NEMS Report Writer to .npz restart file,
        # place the restart files for ref2025 and ref2023 in a location this can grab to generate the 
        # necessary files from ref2023 from AEO2023. Remove this for AEO2026. 
        restartnpzpath = "restartref2025.npz"
        restartnpzpath1 = "restartref2023.npz"
        user.file_map["RestartNPZ"] = restartnpzpath
        dat = dfd = read_npz(user)
        user.file_map["RestartNPZ"] = restartnpzpath1
        dat1 = dfd1 = read_npz(user)
        dfdkeys = pd.DataFrame(dfd.keys())
        dfdkeys1 = pd.DataFrame(dfd1.keys())
        dfdkeys[["commonblock", "varname"]] = dfdkeys[0].str.split("/", expand=True)
        dfdkeys1[["commonblock", "varname"]] = dfdkeys1[0].str.split("/", expand=True)
        duplicate_varsnames = dfdkeys["varname"].duplicated(keep=False)
        duplicate_varsnames1 = dfdkeys1["varname"].duplicated(keep=False)
        dfdkeys[0] = dfdkeys[0].mask(duplicate_varsnames, dfdkeys[0])
        dfdkeys[0] = dfdkeys[0].where(duplicate_varsnames, dfdkeys["varname"])
        dfdkeys1[0] = dfdkeys1[0].mask(duplicate_varsnames1, dfdkeys1[0])
        dfdkeys1[0] = dfdkeys1[0].where(duplicate_varsnames1, dfdkeys1["varname"])
        dfdkeys = dfdkeys.drop(columns=["commonblock", "varname"])
        dfdkeys1 = dfdkeys1.drop(columns=["commonblock", "varname"])
        dfdkeys = dfdkeys.set_index([0], drop=False)
        dfdkeys1 = dfdkeys1.set_index([0], drop=False)
        dfdhold = dfd.copy()
        dfdhold1 = dfd1.copy()

        for i in list(dfd):
            try:
                dfdkeys.loc[i]
            except KeyError:
                parts = i.split("/")
                dfdhold[parts[-1].upper()] = dfdhold.pop(i)

        for i in list(dfd1):
            try:
                dfdkeys1.loc[i]
            except KeyError:
                parts1 = i.split("/")
                dfdhold1[parts1[-1].upper()] = dfdhold1.pop(i)

        dfd = dfdhold.copy()
        dfdorig = dfd
        dfd1 = dfdhold1.copy()
        count = 1
        for key, df in dfd.items():
            try:
                dfd[key] *= 0
            except:
                pass
        #dfd = sum_CarbCap_vars(dfd)
        #dfd.update(dfd1)
        # This code updates sections for EPM that due to variable resizing does not work with AEO2023 data. Remove after AEO2025.
        for key, df in dfd1.items():
            try:
                hold1 = dfd1[key]
                hold = dfd[key]
                if hold1.shape == hold.shape:
                    dfd[key] = dfd1[key]
                elif hold1.shape < hold.shape:
                    if key == "EM_INDY":
                        dfd[key].loc[(slice(1, 16), slice(None)), :] = dfd1[key].loc[(slice(1, 16), slice(None)), :]
                        dfd[key].loc[(slice(17, 19), slice(None)), :] = 0.0
                        dfd[key].loc[(20, slice(None)), :] = dfd1[key].loc[(17,slice(None)),:].values
                    elif key == "EM_ELEC":
                        dfd[key].loc[(slice(1, 7), slice(None)), :] = dfd1[key].loc[(slice(1, 7), slice(None)), :]
                        dfd[key].loc[(8, slice(None)), :] = 0.0
                        dfd[key].loc[(9, slice(None)), :] = dfd1[key].loc[(8, slice(None)), :].values
                else:
                    print(key + ': this variable is resized but okay for AEO2025.')
            except:
                print(key + ": this is a new variable and will be left as zero.")

    # Add table print seetings:
    table_spec = {}
    table_spec["settings"] = settings

    # Add aeo published table number mapping to layin table number
    table_spec["table_mapping"] = table_mapping

    # For Tracy
    table_spec["ftab_dat"] = {}
    table_spec["ftab_dat"]["FYRPRC"] = settings["year_index_real_dol_priceyr"]

    # Add general NEMS info from user
    table_spec["StudyID"] = user.StudyID  # 'AEO2024'
    table_spec["LongTitle"] = user.LongTitle  # 'Annual Energy Outlook 2023'
    table_spec["case_descriptions"] = user.case_descriptions  # 'Reference Case'
    table_spec["ScenLabel"] = user.ScenLabel  # 'Reference'
    table_spec["Scenario"] = user.Scenario  # example: 'ref2024'
    table_spec["DateKey"] = user.DateKey  # example: 'd101323b'
    table_spec["ScenDate"] = user.ScenDate  # 'ref2024.d020623a'
    table_spec["case_descriptions"] = user.case_descriptions  # ?
    table_spec["Release_Date"] = user.release_date

    # Add Projection years from user
    table_spec["years"] = years  # [1990, 1991, ..., 2050]
    table_spec["year_len"] = len(years)
    table_spec["years_str"] = years_str  # ['1990', '1991', ..., '2050']
    table_spec["first_year"] = user.first_year  # 1990
    table_spec["last_year"] = user.last_year  # 2050

    # Add Print years from config.ini
    table_spec["first_print_year"] = first_print_year  # 2020
    table_spec["last_print_year"] = last_print_year  # 2050
    table_spec["print_years"] = print_years  # [2020, 2021, ..., 2050]

    # Add growth rate start to calculate ave annual growth - may differ from first_print_year
    table_spec["growth_rate_start"] = int(settings["growth_rate_start"])

    # Add Print regions flag, 1 - print regions, 0 - not
    table_spec["print_regions"] = int(settings["print_regions"])

    # Add Input file specs from user
    table_spec["table_ids_full"] = table_ids_full
    table_spec["table_ids_req"] = table_ids_req
    table_spec["restartnpz_path"] = restartnpz_path
    table_spec["table_var_def_path"] = table_var_def_path
    table_spec["df_format"] = df_format
    table_spec["table_input_df"] = table_input_df  # more convenient to use
    table_spec["table_input"] = table_input
    table_spec["dat"] = dat
    table_spec["dfd"] = dfd
    table_spec["coefficients_path"] = coefficients_path
    table_spec["layin_ver"] = layin_ver
    table_spec["citation_path"] = user.citation_path
    # table_spec['tabreq_path'] = user.tabreq_path
    table_spec["MACYR"] = MACYR

    # Add RAN required info from user
    table_spec["RSVer"] = user.RSVer  # 'AA01' Scenario Header data
    table_spec["RSFile"] = user.RSFile  # restart.npz
    table_spec["RSScen"] = user.RSScen  # Example: ../../ref2024/d101323b/RESTART.unf
    if outputdir['outputpath'] == 'output':
        table_spec["all_row_file_path"] = os.path.join(
            os.getcwd(), "output", "NEMS" + table_spec["Scenario"] + ".unif.api.csv"
        )
    else:
        table_spec["all_row_file_path"] = os.path.join(
            outputdir['outputpath'], "NEMS" + table_spec["Scenario"] + ".unif.api.csv"
        )

    return table_spec, dat


@log_execution(main_logger, message="fill_base_table")
def fill_base_table(table_spec, table_id, dfd_sme):
    """
    Fill a base table using specified program and variables.

    Parameters
    ----------
    table_spec : dict
        Table specifications (metadata).
    table_id : str
        ID of the table to be filled, e.g., 'TN 002'
    dfd_sme : dict
        Dictionary containing SME (Subject Matter Expert) data.

    Returns
    -------
    dict
        Dictionary containing the filled table rows.

    """

    main_logger.info(f"{table_id}")

    # Load table variable definitions from the TOML file
    with open(table_spec["table_var_def_path"], "rb") as f:
        temp = tomllib.load(f)
    my_vars = temp[table_id]["v"]

    # Add dfd_sme to dfd
    dfd.update(dfd_sme)

    # Get the base program
    my_program = table_spec["table_input"][table_id]["Program"]

    # Fill the table using the specified program and variables
    d_df = globals()[my_program](dfd, table_spec, table_id)

    return d_df


@log_execution(main_logger, message="make_base_tables")
def make_base_tables(table_spec, dfd_sme, user):
    """
    Create base tables.

    Parameters
    ----------
    table_spec : dict
        Table specifications (metadata).

    dfd_sme : dict of DataFrame
        Dataframe containing the SME (Subject Matter Expert) data.

    Returns
    -------
    dict
        Dictionary containing the base tables.
    """

    d_base = {}

    for table_id in table_spec["table_ids_full"]:
        # Check if we need to generate RAN which requires all available tables
        # Otherwise we only need to create the base tables for required tables
        if (int(table_spec["settings"]["create_ran_file"]) != 1) and (
            table_id not in table_spec["table_ids_req"]
        ):
            continue

        d_df = fill_base_table(table_spec, table_id, dfd_sme)

        for row, value in d_df.items():
            if isinstance(value, pd.Series):
                d_df[row] = value.to_frame().T

            d_df[row].index.name = "Region"

            if int(config.get("debugging", "debug_base")) == 1:
                main_logger.info(
                    f"row: {row}, Type: {type(value).__name__}, Value_shape: {value.shape}"
                )

            if len(table_spec["table_input"][table_id]["Region"]) == 1:
                d_df[row].index = [0] * len(d_df[row])
                d_df[row].index.name = "Region"

        d_base[table_id] = pd.concat(
            [df.dropna(axis=1, how="all") for df in d_df.values()], keys=d_df.keys()
        )
        # It's better to leave NaN and inf in the base table to facilitate debugging.
        # We'll replace NaN and inf with 0 in def format_table !!!
        d_base[table_id] = pd.concat([df for df in d_df.values()], keys=d_df.keys())

        # Cut redundant year columns produced by base table programs
        d_base[table_id] = d_base[table_id].iloc[:, : len(table_spec["years"])]

        # Create base table Excel files
        if int(table_spec["settings"]["print_base_tables"]) == 1:
            save_base_tables(d_base, table_spec, table_id, user)

    return d_base


@log_execution(main_logger, message="save_base_tables")
def save_base_tables(d_base, table_spec, table_id, user):
    """
    Save base tables to Excel.

    Parameters
    ----------
    d_base : dict
        Dictionary containing the base tables.
    table_spec : dict
        Table specifications (metadata).
    table_id : str
        ID of the table.

    Returns
    -------
    None
    """

    options = {}
    options["strings_to_formulas"] = False
    options["strings_to_urls"] = False
    kwargs = {"options": options}

    if user.outputdir == "output":
        base_file_path = os.path.join(os.getcwd(), "output", table_id + " base.xlsx")
    else:
        base_file_path = os.path.join(user.outputdir, table_id + " base.xlsx")

    # pylint disable=E0110
    with pd.ExcelWriter(base_file_path, "xlsxwriter", engine_kwargs=kwargs) as writer:
        # Iterate through each Region
        for region, region_name in table_spec["table_input"][table_id][
            "Region"
        ].items():
            df_base_region = d_base[table_id][
                (d_base[table_id].index.get_level_values("Region") == region)
                | (d_base[table_id].index.get_level_values("Region").isna())
            ]

            df_base_region.to_excel(
                writer, sheet_name="Region" + str(region), index=True
            )

    return


@log_execution(main_logger, message="format_table")
def format_table(d_base, table_id, table_spec):
    """Format a base table according to specified formatting rules.

    Parameters
    ----------
    d_base : dict
        Dictionary containing the base tables
    table_id : str
        ID of the table to be formatted
    table_spec : dict
        Table specifications and metadata

    Returns
    -------
    pd.DataFrame
        Formatted table
    """

    main_logger.info(f"{table_id}")

    # Fill NaN with 0 to avoid the entire row data missing!
    df_table = d_base[table_id].fillna(0)

    # Replace inf and -inf with 0
    df_table = df_table.replace([float("inf"), -float("inf")], 0)

    # Drop column 0 if presented
    if 0 in df_table.columns:
        df_table = df_table.drop(0, axis=1)

    # Name indexes
    df_table.index.names = ["IROWS", "Region"]

    # Change IROWS from int to str to deal with layin.csv
    if table_spec["layin_ver"] == "csv":
        df_table.index = df_table.index.set_levels(
            df_table.index.levels[0].astype(str), level=0
        )

    # Remove redundant columns identified in base program 103, 112...
    df_table.columns = table_spec["years"]

    # Get the current table format
    df_format = table_spec["df_format"]
    my_label = df_format.loc[
        df_format["TableNumber"] == int(table_id.split(" ")[-1])
    ].copy()
    merged_dfs = []

    # Iterate through each region
    for region, region_name in table_spec["table_input"][table_id]["Region"].items():

        # Get region data
        df_reg = df_table.loc[(slice(None), region), :]

        # Merge table data with table format
        merged_df = pd.merge(
            df_reg, my_label, left_on="IROWS", right_on="IROWS", how="right"
        ).fillna("")

        merged_df["Region"] = region
        merged_df["RegionNum"] = region

        # Create/Modify uid for data rows by adding RegionNum
        merged_df["uid"] = (
            merged_df["table_name"] + str(region).zfill(3) + ":" + merged_df["VarNam2"]
        )

        merged_df.loc[merged_df["VarNam2"] == "", "uid"] = ""

        # Fill in region names
        # This operation should only be performed for tables with region versions like Table 2
        # Some national tables consist of regional data blocks (e.g., Table 64) with 'Geogr'
        # (original 'Geography Override') filled with region names - should not be overwritten
        if len(table_spec["table_input"][table_id]["Region"]) > 1:
            merged_df["Geogr"] = region_name
        else:
            merged_df["Geogr"] = merged_df["Geography Override"]
            # 'Geography Override' is blank for some tables e.g., table 18 in layin.
            # But the all row csv file needs to fill in with 'United States'
            # Check if the column is entirely NaN or blank
            is_entirely_blank = (
                merged_df["Geogr"].isna().all() or merged_df["Geogr"].eq("").all()
            )
            # Set to 'United States'
            if is_entirely_blank:
                merged_df["Geogr"] = "United States"

        merged_dfs.append(merged_df)

    df_table = pd.concat(merged_dfs, ignore_index=True)

    # Remove rows of CM = 'CM' or 'PB'
    df_table = df_table[~((df_table["CM"] == "CM") | (df_table["CM"] == "PB"))]

    # Check 'Bonus rows'
    # DRForm character 6:
    #   Indicates Bonus Rows - Internal Only in published tables.
    #       0: regular publishable row
    #       1: line is omitted in formal tables and xml output files
    # Bonus rows are printed to the Excel reports only if print_bonus_rows = 1
    if int(table_spec["settings"]["print_bonus_rows"]) != 1:
        # Remove 'Bonus' rows - DRForm character 6 = 1
        df_table = df_table[df_table["DRForm"].str[-1] != "1"]

    # Set the original indexes back
    df_table.set_index(["IROWS", "Region"], inplace=True)

    # Move uid and Label columns to the begining
    cols = df_table.columns.tolist()
    cols = ["uid", "Label"] + [col for col in cols if col not in ["uid", "Label"]]
    df_table = df_table[cols]

    # # Swap the positions of the index levels
    # df_table = df_table.swaplevel('Row', 'Region')

    # # Sort the DataFrame by the first index level
    # df_table = df_table.sort_index(level='Region')

    df_table.fillna("", inplace=True)

    return df_table


@log_execution(main_logger, message="make_formatted_tables")
def make_formatted_tables(table_spec, d_base):
    """Creates formatted tables based on table spec and base tables.

    Parameters
    ----------
    table_spec : dict
        Table specifications and metadata.
    d_base : dict
        Dictionary containing the base tables.

    Returns
    -------
    dict
        Dictionary containing the formatted tables.
    """

    d_formatted = {}

    for table_id in list(table_spec["table_input"].keys()):  # table: 'TN 001', ...
        # We need created formatted tables of all base tables if RAN is required !
        # Otherwise just create formatted tables for required tables
        if (int(table_spec["settings"]["create_ran_file"]) != 1) and (
            table_id not in table_spec["table_ids_req"]
        ):
            continue

        df = format_table(d_base, table_id, table_spec)

        # Set region numbers in published tables --------------------
        if len(table_spec["table_input"][table_id]["Region"]) > 1:
            if len(table_spec["table_input"][table_id]["Region"]) == 11:
                # Census region: 11->0
                df.reset_index(inplace=True)
                df.loc[df["Region"] == 11, "Region"] = 0
                df.loc["RegionNum"] = df["Region"]

                # remove 10
                df.drop(df[df["Region"] == 10].index, inplace=True)

            elif len(table_spec["table_input"][table_id]["Region"]) == 28:
                # Census region: 28 ->0,
                df.reset_index(inplace=True)
                df.loc[df["Region"] == 28, "Region"] = 0

                # remove 26, 27
                df.drop(df[df["Region"] == 26].index, inplace=True)
                df.drop(df[df["Region"] == 27].index, inplace=True)
            else:
                pass  # any other regions?

            # modify uid, e.g., QUA011:ca_LiquefiedPetr -> QUA000:ca_LiquefiedPetr
            a = df.loc[(df["Region"] == 0) & (df["uid"].str.strip() != "")]
            a.loc[:, "uid"] = a["uid"].str[:4] + "00" + a["uid"].str[6:]
            df.loc[a.index, "uid"] = a["uid"]

            # Reset index
            df.set_index(["IROWS", "Region"], inplace=True)

        else:
            pass  # do nothing for tables with national version only
        # -----------------------------------------------------------------------

        # Calculate Average Annual Change during print years (%)
        # TODO: check with EIA for the rate calculation method
        #   '' errors occured due to empty or non-numeric string values in the columns
        val_last_y = pd.to_numeric(df[table_spec["last_print_year"]], errors="coerce")
        # val_first_y = pd.to_numeric(df[table_spec['first_print_year']], errors='coerce')
        val_first_y = pd.to_numeric(
            df[table_spec["growth_rate_start"]], errors="coerce"
        )

        # avg_annual_change = (val_last_y - val_first_y) / (len(table_spec['print_years'])-1)4
        # Where there is a zero in first year or last year, do not do growth rate change. Instead
        # replace with nan so we can fill nan with blank.
        val_last_y.replace(0,np.nan,inplace=True)
        val_first_y.replace(0,np.nan,inplace=True)
        avg_annual_change = (val_last_y / val_first_y) ** (
            1 / (len(table_spec["print_years"]) - 1)
        ) - 1

        avg_annual_change = avg_annual_change.fillna("")

        df["Avg Annual Change"] = (
            avg_annual_change  # will be formatted in def add_style
        )

        d_formatted[table_id] = df.copy()

    return d_formatted


@log_execution(main_logger, message="make_all_row_table")
def make_all_row_table(table_spec, d_formatted, user):
    """Create a table containing all rows from all formatted tables.

    Parameters
    ----------
    table_spec : dict
        Table specifications (metadata).
    d_formatted : dict
        Dictionary containing the formatted tables.

    Returns
    -------
    pd.DataFrame
        Table containing all rows from all formatted tables.

    """

    l_df = []

    # Specify columns for the all row csv file like NEMSref2024.unif.api)
    cols_csv = [
        "TableNumber",
        "RowNum",
        "RegionNum",
        "VarName",
        "VarNam2",
        "GLabel",
        "Gunits",
        "RowFmt",
        "DaType",
        "SubDat",
        "Sector",
        "SubSec",
        "Source",
        "SubSrc",
        "Geogr",
    ]

    cols_csv += table_spec["years"]

    for table_id in table_spec["table_ids_full"]:
        # We need all formatted tables if RAN is required !
        # Otherwise just include the formatted tables for required tables
        if (int(table_spec["settings"]["create_ran_file"]) != 1) and (
            table_id not in table_spec["table_ids_req"]
        ):
            continue

        df = d_formatted[table_id].copy()
        df = df[
            df.index.get_level_values("IROWS").astype(str).str.isdigit()
            & ~df.index.get_level_values("IROWS").isna()
        ]

        df = df[cols_csv]
        l_df.append(df)

    dfs = pd.concat(l_df)

    # Insert 'Scenario' column at index 0
    dfs.insert(0, "Scenario", table_spec["Scenario"])

    # Insert 'Datakey' column at index 1
    dfs.insert(1, "Datekey", table_spec["DateKey"])

    # Reset region number
    dfs.reset_index(inplace=True)
    dfs["RegionNum"] = dfs["Region"]
    dfs.drop(["IROWS", "Region"], inplace=True, axis=1)
    dfs.sort_values(["TableNumber", "RegionNum"], inplace=True)

    # Force the conversion of year columns to float data type
    #   '' errors occured due to empty or non-numeric string values in the columns
    for year in table_spec["years"]:
        dfs[year] = pd.to_numeric(dfs[year], errors="coerce")

    # Create all row csv table file ---
    #  e.g., 'NEMSref2024.unif.api.csv'
    all_row_file_path = table_spec["all_row_file_path"]

    if not os.path.exists(user.outputdir):
        os.makedirs(user.outputdir)

    dfs.to_csv(all_row_file_path, float_format="%.6f")

    return dfs


# @log_execution(main_logger, message="add_style")
def add_style(workbook, sheet, df_all, col_num, stacked=False):
    """Applies formatting style to each row and in the sheet in font size, plain/bold, color,...

       layin:
        F.  DRForm specifies a 6-character string to control formatting, as follows:
            e.g., p15000

        position 1:  x, y, p, or b or B, i or I
                    x: Skip this row in formal tables only (blank row used for spacing)
                    y: skip this row always
                    p: plain font (not bolded)
                    B: bold font
                    i or I: italic font
        position 2:  Indent level for row label.
                    0: zero or no indent
                    1: 1st indent (1 character or 1/8th inch)
                    2. 2nd indent (2 characters or 2/8ths inch)
                    3. 3rd indent (3 characters of 3/8ths inch)
                    4. 4th indent etc
        character 3:  Font size.  5: normal.  2: 70%, 3: 80% 4: 90%, 5: 111% 6: 125%  7: 143%
        character 4:  not used -> color !
        character 5:  Attempting to use this position (set to '1') denoting row in table that
                      table browser opens to. If no rows designated in table, browser defaults to first row.
        character 6:  Indicates Bonus Rows - Internal Only in published tables.
                    0: regular publishable row
                    1: line is omitted in formal tables and xml output files

    Parameters
    ----------
    sheet : object (xlsxwriter worksheet)
        The sheet to apply formatting to.
    df_all : DataFrame
        The DataFrame containing the data to be formatted.
    col_num : int
        The number of columns in the sheet (reporting years + Avg Annual Change).

    Returns
    -------
    object (xlsxwriter worksheet)
        The modified sheet with 'style'

    Note
    ----
    character 4 was not used in FTab. We use it to specify color ->
        r: red, g: green, b: blue,
        (0: default to black)

    """

    # Get font style: character 1
    df_all["bold"] = df_all["DRForm"].str.startswith("B")
    df_all["italic"] = df_all["DRForm"].str.startswith(("I", "i"))

    # Get font size: character 3
    sizes = {
        "0": 9.0,  # Code 0: default
        "2": 6.0,  # Code 2 corresponds to 70%
        "3": 7.0,  # Code 3 corresponds to 80%
        "4": 8.0,  # Code 4 corresponds to 90%
        "5": 9.0,  # Code 5 corresponds to 111%; default
        "6": 10.0,  # Code 6 corresponds to 125%
        "7": 12.0,  # Code 7 corresponds to 143%
    }

    df_all["font_size"] = df_all["DRForm"].str[2].fillna("5").replace(sizes)

    # Get color: character 4
    colors = {
        "0": "#000000",  # black, default
        "1": "#000000",  # black, default
        "2": "#000000",  # black, default
        "8": "#000000",  # black, default
        "r": "#FF0000",  # red
        "g": "#00B050",  # green
        "b": "#0066CC",  # blue '4472C4'
    }
    df_all["font_color"] = df_all["DRForm"].str[3].fillna("0").replace(colors)

    # Property columns to get from rows of DataFrame
    props = ["bold", "italic", "font_color", "font_size"]

    # Format for thick blue bottom border
    bdrfmt = workbook.add_format({"bottom_color": "#0096D7", "bottom": 5})

    # Apply formatting to each row and cell
    # NOTE: xlsx writer is 0-indexed
    for i, row in enumerate(df_all.fillna("").to_dict("records")):
        fmtd = {prop: row[prop] for prop in props}
        if i == 0:  # First row formatting
            fmtd["bold"] = True
            fmtd["italic"] = False
            fmtd["num_format"] = "yyyy"
            if not stacked:
                fmtd["font_size"] = 12

        try:  # Number type
            row_format = int(row["RowFmt"])
        except ValueError:  # For alphabetical values
            row_format = 9

        fmtd["num_format"] = "#,##0" if row_format == 9 else "#,##0." + "0" * row_format

        # Format RL row  - bold font, and blue bottom side line
        if row["CM"] == "RH":
            fmtd["bold"] = True
            fmtd["font_size"] = 9
            fmtd["num_format"] = "0"
            sheet.conditional_format(
                i, 0, i, col_num - 1, {"type": "no_blanks", "format": bdrfmt}
            )

        fmt = workbook.add_format(fmtd)
        sheet.set_row(i, cell_format=fmt)
        fmtd["num_format"] = "0.0%"
        pctfmt = workbook.add_format(fmtd)
        sheet.conditional_format(
            i,
            col_num - 1,
            i,
            col_num - 1,
            {"type": "no_blanks", "format": pctfmt},
        )
    # END row formatting

    # Set last column as % change
    sheet.set_column(col_num - 1, col_num - 1, options={"num_format": "0.0%"})

    # Add blue bottom side of the last RL row(s)
    # Find the last "RL" row with  by TableNumber (may have multi-tables in stack)
    bdrrows = (
        df_all.loc[df_all["CM"] == "RL"]
        .reset_index()
        .groupby(["table_name", "RegionNum"])["index"]
        .max()
        .to_list()
    )
    for row in bdrrows:
        for col in range(1, col_num):
            sheet.write_blank(row - 1, col, "", bdrfmt)

    # Hide first column, set second column width, freeze panes
    sheet.set_column("A:A", options={"hidden": True})
    sheet.set_column("B:B", 40)
    sheet.freeze_panes("C2")

    return sheet


@log_execution(main_logger, message="stack_tables")
def stack_tables(
    df_stack,
    d_stack,
    cols_final,
    print_years,
    ScenDate,
    file_name,
    link_start_row,
    link_start_col,
    user,
):
    """Stacks tables in a DataFrame and generates an Excel file.

    Parameters
    ----------
    df_stack : pandas.DataFrame
        DataFrame containing the stacked tables.
    d_stack : dict
        Dictionary mapping table captions to the first row index in the stacked tables.
    cols_final : list
        List of column names to include in the stacked tables.
    print_years : list
        List of years of the report.
    ScenDate : str
        Scenario name and date for sheet name of the stacked tables.
    file_name : str
        Name of the output Excel file.
    link_start_row : int
        The row number from which the hyperlink bookmarks should start.
    link_start_col : int
        The column number from which the hyperlink bookmarks should start.

    Returns
    -------
    None
        The function saves the stacked tables DataFrame to an Excel file.
    """

    main_logger.info("\n============ stack_tables ============")
    if user.outputdir == 'output':
        formatted_file_path = os.path.join(os.getcwd(), "output", file_name)
    else:
        formatted_file_path = os.path.join(user.outputdir, file_name)
    sheet_name = ScenDate

    # pylint: disable=E0110
    with pd.ExcelWriter(formatted_file_path, engine="xlsxwriter") as writer:
        # Create a workbook and a stack sheet
        workbook = writer.book
        sheet = workbook.add_worksheet(sheet_name)

        # Add style: font size, plain/bold, color,...
        sheet = add_style(workbook, sheet, df_stack, 2 + len(print_years) + 1, True)

        # Clean sheet columns
        df_out = df_stack[cols_final]
        df_out.to_excel(writer, sheet_name=sheet_name, header=False, index=False)

        # Create bookmark links at column C, starting from row 10
        fmt = workbook.add_format(
            {"underline": 1, "font_color": "#0000FF", "bold": False, "italic": False}
        )
        for i, (name, first_row) in enumerate(d_stack.items()):
            sheet.write_url(
                row=i + link_start_row,
                col=link_start_col - 1,  # Column C, 0-indexed
                url=f"#'{sheet.get_name()}'!B{first_row}",
                cell_format=fmt,  # Apply hyperlink style
                string=name,
            )
        fmtd = {
            "font_size": 9,
            "bold": True,
            "bottom_color": "#0096D7",
            "bottom": 5,
        }
        hdrfmt = workbook.add_format(fmtd)
        sheet.set_row(0, cell_format=hdrfmt)


@log_execution(main_logger, message="write_tables")
def write_tables(
    table_spec, d_formatted, table_nums, file_name, region_flg, stack, map_aeo_table, user,
):
    """
    Write formatted tables with style to Excel files using xlsxwriter engine.

    Parameters
    ----------
    table_spec : dict
        Table specifications and metadata.
    d_formatted : dict
        Dictionary containing the formatted tables.
    table_nums : list
        List of table numbers to be processed [1, 2, 3,...].
    file_name : str
        Name of the output Excel file.
    region_flg : int
        Flag indicating whether regional tables are required (1 for yes, 0 for no).
    map_aeo_table: True/False

    Note
    ----
    Writing base table was moved to save_base_tables.
    """

    main_logger.info("\n============ write_tables ============")

    write_table_successful = False

    # Get table id list
    table_ids = ["TN " + str(num).zfill(3) for num in table_nums]

    # Check if all tables in table_nums are available
    my_table_input = {k: table_spec["table_input"][k] for k in table_ids}

    # Define the final formatted table columns
    cols_final = ["uid", "Label"] + table_spec["print_years"] + ["Avg Annual Change"]

    # Create a bookmark dict storing table captions and locations
    d_stack = {}  # {table caption: first row in df_stack}
    link_start_row = 10
    link_start_col = 3  # sheet column C

    # Calculate the total table numbers including regions
    if region_flg == 1:
        table_count = sum(
            len(table_dict["Region"]) for table_dict in my_table_input.values()
        )
    else:
        table_count = len(my_table_input)

    # Create a df to hold all stacked tables with 10 blank rows to hold table info
    df_stack = pd.DataFrame(index=range(table_count + 10))

    # Make sure tables are required in tabreq.txt !!!
    for table in my_table_input:
        main_logger.info("table: %s", table)

        # Check if the table is indeicated as required in tabreq.txt:
        #   e.g., 45 is a stacked table in sup_tran.xlsx, but may be indicated as not required.
        #   In that cae, it will not appear in sup_tran.xlsx
        if table not in table_spec["table_ids_req"]:
            main_logger.warning(
                "table: %s is not indicated as required in tabreq !", table
            )
            continue

        table_num = int(table.split(" ")[1])

        # Mapping layin table num to aeo published table num '''
        if map_aeo_table:
            try:
                table_num_print = (
                    table_spec["table_mapping"][table_num] if map_aeo_table else table_num
            )
            except KeyError:
                continue
        else:
            table_num_print = (
                    table_spec["table_mapping"][table_num] if map_aeo_table else table_num
            )
                

        # Create a writer for the table
        if user.outputdir == "output":
            formatted_file_path = os.path.join(os.getcwd(), "output", file_name)
        else:
            formatted_file_path = os.path.join(user.outputdir, file_name)
        with pd.ExcelWriter(
            formatted_file_path, engine="xlsxwriter"
        ) as writer:  # pylint: disable=E0110
            # Set region numbers in published tables
            region_num_aeo = (
                range(0, 9 + 1)
                if len(my_table_input[table]["Region"]) == 11
                else (
                    range(0, 25 + 1)
                    if len(my_table_input[table]["Region"]) == 28
                    else range(0, 0 + 1)
                )
            )

            # Get table format
            df_format = table_spec["df_format"]

            df_format = df_format.loc[
                df_format["TableNumber"] == table_num,
                ["TableNumber", "CM", "Label", "IROWS", "RowNum", "RowFmt"],
            ]
            df_format = df_format.reset_index(drop=True)

            # Check if regional tables required (set to [0] if national table only)
            region_num_aeo = region_num_aeo if region_flg == 1 else [0]

            # Remove non print years, e.g, 1990-1999
            df_data_tab = d_formatted[table]

            col_list = [
                col
                for col in df_data_tab.columns
                if col
                not in range(
                    int(table_spec["first_year"]), int(table_spec["first_print_year"])
                )
            ]
            df_data_tab = df_data_tab[col_list]

            # Iterate through each region
            # for region, region_name in my_table_input[table]['Region'].items():
            for region_num in region_num_aeo:
                df_data_reg = df_data_tab[
                    (df_data_tab.index.get_level_values("Region") == region_num)
                    | (df_data_tab.index.get_level_values("Region").isna())
                ].copy()
                df_data_reg.reset_index(drop=True, inplace=True)

                # Merge data and format
                df = pd.merge(
                    df_data_reg,
                    df_format,
                    left_index=True,
                    right_index=True,
                    how="right",
                ).fillna("")
                df = df.rename(
                    columns={"Label_x": "Label", "CM_x": "CM", "RowFmt_y": "RowFmt"}
                )
                df = df.reset_index(drop=True)

                # df structure example:
                # -----------------------------------------------------------------------------------
                # index   uid                       Label                                    1990,...
                #     0                             SUP
                #     1
                #     2                             Total Energy Supply, Disposition,...
                #     3                             (quadrillion Btu, unless otherwise noted)
                #     4                             Supply, Disposition, and Prices
                #     5
                #     6                             Production
                #     7   SUP000:ba_CrudeOilLease     Crude Oil and Lease Condensate
                # ...
                # -----------------------------------------------------------------------------------

                # Target table structure & style (refer to EIA xml file)
                # ---------------------------------------------------------------------------
                #   1 HD1: bold, 12, blue
                #   2 HD2: plain
                #   3 HD3: (blank)
                #   4 RH:  bold [row head label + 1990...2050]
                #   5 (blank)
                #   ...
                #
                # Default format style:
                #   font: plain, size: 9, color: black
                # ---------------------------------------------------------------------------

                # HD rows, e.g., Energy Consumption by Sector and Source
                df_HD = df.loc[df["CM"] == "HD"].copy()

                # Customize table captions
                if region_num == 0:
                    # National table: 2. Energy Consumption by Sector and Source
                    table_caption = f"{table_num_print}. {df_HD.iloc[0]['Label']}"
                else:
                    # Regional table: 2.1. Energy Consumption by Sector and Source
                    table_caption = (
                        f"{table_num_print}.{region_num}. {df_HD.iloc[0]['Label']}"
                    )
                    # Add region num and name to HD3
                    df_HD.iloc[2, 1] = (
                        f"{str(region_num).zfill(2)} - {my_table_input[table]['Region'][region_num]}"
                    )

                df_HD.iloc[0, 1] = table_caption

                # Create a blank row between HD and RH
                df_blk = pd.DataFrame([[""] * len(df.columns)], columns=df.columns)
                df_blk["Label"] = ""

                # RH row: RH Label + [1990-2050] + ['Avg Annual Change']
                df_RH = df.loc[df["CM"] == "RH"].copy()
                df_RH.loc[:, table_spec["print_years"]] = table_spec["print_years"]
                columnsize = 37 + len(table_spec["print_years"])
                new_rows_header = pd.DataFrame([[""]*columnsize]*6, columns=df_HD.columns)
                df_HD = pd.concat([new_rows_header, df_HD], ignore_index=True)
                df_HD.iloc[0,1] = table_spec["Scenario"] +"." + table_spec["DateKey"]

                df_HD.iloc[2, 2] = "Report"
                #MAKE THIS PULL FROM SOMEWHERE THAT IT IS 2025, NOT USING SCENARIO
                df_HD.iloc[2, 3] = (
                    table_spec["LongTitle"]
                )  # ref2024 -> 2024
                df_HD.iloc[3, 2] = "Scenario"
                df_HD.iloc[3, 3] = table_spec["Scenario"]  # ref2024
                df_HD.iloc[3, 5] = table_spec["ScenLabel"]
                #ADD SOME MAPPING HERE TO GET OUT SHORT SCENARIO NAME

                df_HD.iloc[4, 2] = "Datekey"
                df_HD.iloc[4, 3] = table_spec["DateKey"]  # ref2024 -> 2024

                df_HD.iloc[5, 2] = "Release Date"
                #CHANGE THIS MAPPING HERE TO SOMEWHERE WE CAN HAVE IT STATE RELEASE MONTH/DAY.
                df_HD.iloc[5, 4] = table_spec["Release_Date"]
                df_HD.loc[df_HD.index[-1], ["Avg Annual Change"]] = "Avg Annual Change"
                df_RH.loc[:, ["Avg Annual Change"]] = (
                    f"{table_spec['growth_rate_start']}-{table_spec['last_print_year']}"
                )

                # Remaining rows/ RL + FN
                df_rst = df.loc[(df["CM"] == "RL") | (df["CM"] == "FN")].copy()
                # Replace any blank cells, inf/-inf cells with '--'
                df_rst_filtered = df_rst[(df_rst['uid'] != '') & (df_rst['Label'] != "") & (df_rst[2024] != "")]
                df_rst_filtered.replace(to_replace=[np.nan, np.inf, -np.inf, ""], value="--",regex=True, inplace=True)
                df_rst.loc[(df_rst['uid'] != '') & (df_rst['Label'] != "") & (df_rst[2024] != ""),:] = df_rst_filtered
                # Add logic to check first year, last year. If value is zero- replace growth with "--".

                # Create table with desired structure
                df_all = pd.concat([df_HD, df_RH, df_rst], ignore_index=True)

                # Create a df to stack all tables (including regions) in the dict
                if stack:
                    df_stack = pd.concat(
                        [df_stack, df_blk, df_blk, df_blk, df_all], ignore_index=True
                    )

                    # Update table row location for bookmark link
                    #   e.g., {'Total Energy Supply, Disposition, and Price Summary': 161}
                    d_stack.update(
                        {table_caption: df_stack.shape[0] - df_all.shape[0] + 1}
                    )

                    # Update bookmark link cell value
                    row_num = len(d_stack)
                    df_stack.iloc[row_num + link_start_row - 1, link_start_col - 1] = (
                        table_caption
                    )

                # Make df index "1" based to easily align with Excel row number
                df_all.index += 1

                # Create a workbook and a region sheet
                workbook = writer.book
                sheet_name = f"Region{region_num if (region_flg == 1 and region_num >0) else table_spec['ScenDate']}"
                sheet_name = table_spec["ScenDate"]
                if region_flg == 1:
                    if region_num > 0:
                        sheet_name = sheet_name + "." + str(region_num)
                    else:
                        continue
                sheet = workbook.add_worksheet(sheet_name)

                # Add style: font size, plain/bold, color, etc.
                col_num = len(cols_final)
                sheet = add_style(workbook, sheet, df_all, col_num)

                # Clean sheet columns
                df_all = df_all[cols_final]

                df_all.to_excel(
                    writer, sheet_name=sheet_name, header=False, index=False
                )

        write_table_successful = True

    # Write stacked table ---
    #   Stacked table top rows:
    # '''--------------------------------------------------------------------
    #             B           C               D
    # -----------------------------------------------------------------------
    # 1   ref2023.d020623a	2022
    # 2
    # 3	                    Report	        Annual Energy Outlook 2023
    # 4	                    Scenario	    ref2023
    # 5	                    Datekey	        d020623a
    # 6	                    Release Date    March 2023
    # 7
    # 8
    # 8
    # 10	Click on link below to go to table.  Press Ctrl+Home to return here.
    # 11	1. Total Energy Supply, Disposition, and Price Summary
    # 12	2. Energy Consumption by Sector and Source
    # 13  ......
    # -----------------------------------------------------------------------'''
    if stack:
        # In case the table is not required in tabreq
        if not df_stack.isnull().all().all():

            # Add instruction plus years in the first row
            df_stack.loc[0, table_spec["print_years"]] = table_spec["print_years"]

            df_stack.iloc[0, 1] = table_spec["ScenDate"]  # ref2023.d020623a

            df_stack.iloc[2, 2] = "Report"
            df_stack.iloc[2, 3] = (
                table_spec["LongTitle"]
            )  # ref2024 -> 2024

            df_stack.iloc[3, 2] = "Scenario"
            df_stack.iloc[3, 3] = table_spec["Scenario"]  # ref2024
            df_stack.iloc[3, 5] = table_spec["ScenLabel"]


            df_stack.iloc[4, 2] = "Datekey"
            df_stack.iloc[4, 3] = table_spec["DateKey"]  # ref2024 -> 2024

            df_stack.iloc[5, 2] = "Release Date"
            df_stack.iloc[5, 4] = table_spec["Release_Date"]

            #IF STACK?
            df_stack.iloc[9, 2] = (
                "Click on link below to go to table.  Press Ctrl+Home to return here."
            )
            #IF STACK?
            # Make df index "1" based to easily align with Excel row number
            df_stack.index = df_stack.index + 1

            # Create Excel file stacking all required tables
            stack_tables(
                df_stack,
                d_stack,
                cols_final,
                table_spec["print_years"],
                table_spec["ScenDate"],
                file_name,
                link_start_row,
                link_start_col,
                user,
            )

    return write_table_successful


@log_execution(main_logger, message="split_region")
def split_region(file_name, ScenDate, user):
    """Splits an Excel file into individual worksheets and saves them as separate files.

    Parameters
    ----------
    file_name : str
        The name of the input Excel file to be split.

    Returns
    -------
    None
        The function does not return any value, but it creates new Excel files
        in the 'output' directory.
    """

    # Load the Excel file
    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name)
    else:
        file_path = os.path.join(user.outputdir, file_name)
    workbook = load_workbook(file_path, keep_vba=True, data_only=False)

    # Loop through each worksheet
    # for sheet_name in workbook.sheetnames:
    for sheet_name in workbook.sheetnames[0:]:  # skip the national sheet
        # Get the worksheet
        worksheet = workbook[sheet_name]

        # Create the output file name
        if sheet_name[-2] == '.':
            regname = sheet_name[-2:]
        else:
            regname = sheet_name[-3:]
        output_file = f"{file_name.split('.')[0][:-3] + regname}.xlsx"
        #output_file = f"{file_name.split('.')[0] + '.' + sheet_name[6:]}.xlsx"
        if user.outputdir == "output":
            output_file_path = os.path.join(os.getcwd(), "output", output_file)
        else:
            output_file_path = os.path.join(user.outputdir, output_file)

        main_logger.info("\n   ------------ " + output_file + " ------------")

        # Create a new workbook and copy the worksheet
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        # new_worksheet.title = sheet_name
        new_worksheet.title = ScenDate

        # Copy the data and formatting
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                # Copy the cell value
                new_worksheet.cell(
                    row=row, column=col, value=worksheet.cell(row=row, column=col).value
                )

                # Copy the cell formatting
                cell = new_worksheet.cell(row=row, column=col)
                cell.font = Font(
                    name=worksheet.cell(row=row, column=col).font.name,
                    size=worksheet.cell(row=row, column=col).font.size,
                    bold=worksheet.cell(row=row, column=col).font.bold,
                    color=worksheet.cell(row=row, column=col).font.color,
                )
                cell.border = Border(
                    bottom=worksheet.cell(row=row, column=col).border.bottom
                )
                cell.number_format = worksheet.cell(row=row, column=col).number_format

        # Hide column 'A'
        new_worksheet.column_dimensions["A"].hidden = True  # .visible = False

        # Freeze the pane at cell C2
        new_worksheet.freeze_panes = "C2"

        new_worksheet.column_dimensions["B"].width = 40

        # Save the new workbook
        new_workbook.save(output_file_path)

    return


@log_execution(main_logger, message="create_aeo_files")
def create_aeo_files(table_spec, d_formatted, user):
    """
    Generates various Excel files with format/style of EIA desired.

    Parameters
    ----------
    table_spec : dict
        A dictionary containing the table specification - see def init()
    d_formatted : dict
        A dictionary containing the formatted table data.

    Returns
    -------
    None
        This function does not return any value, but it can createsthe following Excel files:
        - 20 aeo tables (national tables)
        - One yearbyyear file (stacked national tables)
        - Five supplemental files with national and regional tables
        - 95 regional tables (currently seven region embeded files)
        - Any tables specified in tabreq
        _ Base tables for any tables specified in tabreq
    """

    if int(table_spec["settings"]["print_aeotab1_20"]):
        main_logger.info("\n============ aeotab1.xlsx-aeotab20.xlsx ============")
        # aeotab1-20 (national)
        # Table nums here are for layin which will be mapped to published table nums in def write_tables
        d_aeotab = {
            "aeotab1.xlsx": [1],
            "aeotab2.xlsx": [2],
            "aeotab3.xlsx": [3],
            "aeotab4.xlsx": [4],
            "aeotab5.xlsx": [5],
            "aeotab6.xlsx": [6],
            "aeotab7.xlsx": [7],
            "aeotab8.xlsx": [8],
            "aeotab9.xlsx": [9],
            "aeotab10.xlsx": [10],
            "aeotab11.xlsx": [11],
            "aeotab12.xlsx": [12],
            "aeotab13.xlsx": [13],
            "aeotab14.xlsx": [14],
            "aeotab15.xlsx": [15],
            "aeotab16.xlsx": [16],
            "aeotab17.xlsx": [24],
            "aeotab18.xlsx": [17],
            "aeotab19.xlsx": [23],
            "aeotab20.xlsx": [18],
            "suptab_21.xlsx":[30],
            "suptab_22.xlsx":[32],
            "suptab_23.xlsx":[34],
            "suptab_24.xlsx":[35],
            "suptab_25.xlsx":[36],
            "suptab_26.xlsx":[37],
            "suptab_27.xlsx":[38],
            "suptab_28.xlsx":[39],
            "suptab_29.xlsx":[40],
            "suptab_30.xlsx":[41],
            "suptab_31.xlsx":[42],
            "suptab_32.xlsx":[139],
            "suptab_33.xlsx":[140],
            "suptab_34.xlsx":[43],
            "suptab_35.xlsx":[45],
            "suptab_36.xlsx":[46],
            "suptab_37.xlsx":[47],
            "suptab_38.xlsx":[48],
            "suptab_39.xlsx":[49],
            "suptab_40.xlsx":[50],
            "suptab_41.xlsx":[51],
            "suptab_42.xlsx":[52],
            "suptab_43.xlsx":[53],
            "suptab_44.xlsx":[54],
            "suptab_45.xlsx":[55],
            "suptab_46.xlsx":[56],
            "suptab_47.xlsx":[57],
            "suptab_48.xlsx":[148],
            "suptab_49.xlsx":[58],
            "suptab_50.xlsx":[60],
            "suptab_51.xlsx":[113],
            "suptab_52.xlsx":[114],
            "suptab_53.xlsx":[115],
            "suptab_54.xlsx":[62],
            "suptab_55.xlsx":[123],
            "suptab_56.xlsx":[67],
            "suptab_57.xlsx":[70],
            "suptab_58.xlsx":[71],
            "suptab_59.xlsx":[72],
            "suptab_60.xlsx":[81],
            "suptab_61.xlsx":[76],
            "suptab_62.xlsx":[77],
            "suptab_63.xlsx":[78],
            "suptab_64.xlsx":[90],
            "suptab_65.xlsx":[94],
            "suptab_66.xlsx":[95],
            "suptab_67.xlsx":[99],
            "suptab_68.xlsx":[100],
            "suptab_69.xlsx":[20],
            "suptab_70.xlsx":[22],
            "suptab_71.xlsx":[69],
            "suptab_72.xlsx":[133],
            "suptab_73.xlsx":[73],
        }
        for file_name, table_nums in d_aeotab.items():
            write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                region_flg=0,
                stack=False,
                map_aeo_table=True,
                user=user,
            )

    if int(table_spec["settings"]["print_yearbyyear"]):
        main_logger.info("\n============ yearbyyear.xlsx ============")
        # yearbyyear (stacked national)
        # Table nums here are for layin which will be mapped to published table nums in def write_tables
        d_yearbyyear = {
            "yearbyyear.xlsx": [
                1,
                2,
                3,
                4,
                5,
                6,
                7,
                8,
                9,
                10,
                11,
                12,
                13,
                14,
                15,
                16,
                24,
                17,
                23,
                18,
            ]
        }
        for file_name, table_nums in d_yearbyyear.items():
            write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                region_flg=0,
                stack=True,
                map_aeo_table=True,
                user=user,
            )

    if int(table_spec["settings"]["print_supplemental_files"]):
        main_logger.info("\n============ Supplemental files ============")
        # Supplemental files (stacked national & regional)
        # Table nums here are for layin which will be mapped to published table nums in def write_tables
        d_suptab = {
            "sup_elec.xlsx": [62, 123, 67],
            "sup_ogc.xlsx": [70, 71, 72, 81, 76, 77, 78, 90, 94, 95, 99, 100, 20],
            "sup_rci.xlsx": [30, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42, 139, 140, 43],
            "sup_t2t3.xlsx": [2, 3, 17],
            'sup_tran.xlsx': [45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 148, 58, 60, 113, 114, 115],
            "sup_co2.xlsx": [17, 22, 69, 133]
        }
        for file_name, table_nums in d_suptab.items():
            main_logger.info("\n   ------------ %s ------------", file_name)
            write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                table_spec["print_regions"],
                stack=True,
                map_aeo_table=True,
                user=user,
            )

    if int(table_spec["settings"]["print_regional_tables"]):
        main_logger.info("\n============ Regional Tables ============")
        # Regional Tables
        #   EIA proposed naming convention: suptab[table#].[region#].xlsx. But this will generate about 100 Excel files!
        #   It seems more reasonable to put all regional dat in the same table file but in different sheets?
        #
        # Table nums here are for layin which will be mapped to published table nums in def write_tables
        d_regtab = {
            "suptab_2all.xlsx": [2],
            "suptab_3all.xlsx": [3],
            "suptab_18all.xlsx": [17],
            "suptab_38all.xlsx": [48],
            "suptab_54all.xlsx": [62],
            "suptab_56all.xlsx": [67],
            "suptab_57all.xlsx": [70],
            "suptab_73all.xlsx": [73],
        }
        for file_name, table_nums in d_regtab.items():
            main_logger.info("\n   ------------ " + file_name + " ------------")
            write_table_successful = write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                region_flg=1,
                stack=False,
                map_aeo_table=True,
                user=user,
            )
            if write_table_successful:
                split_region(file_name, table_spec["ScenDate"], user)

    if int(table_spec["settings"]["print_tabreq_tables"]):
        # Create the dict of table file names and table num for Tracy testing base tables
        main_logger.info("\n============ print_tabreq_tables ============")
        d_aeotab = {}
        table_nums_req = [int(item[-3:]) for item in table_spec["table_ids_req"]]
        for i, value in enumerate(table_nums_req, start=1):
            file_name = f"TN {str(value).zfill(3)}.xlsx"
            d_aeotab[file_name] = [value]
        for file_name, table_nums in d_aeotab.items():
            write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                table_spec["print_regions"],
                stack=False,
                map_aeo_table=False,
                user=user,
            )

    if int(table_spec["settings"]["print_tabreq_tables_stack"]):
        #  Create the dict of table file names and table num for Tracy testing base tables
        table_nums_req = [int(item[-3:]) for item in table_spec["table_ids_req"]]
        d_stacked = {"all_tables_stacked.xlsx": table_nums_req}
        for file_name, table_nums in d_stacked.items():
            write_tables(
                table_spec,
                d_formatted,
                table_nums,
                file_name,
                table_spec["print_regions"],
                stack=True,
                map_aeo_table=False,
                user=user,
            )


@log_execution(main_logger, message="write_longthin")
def write_longthin(table_spec, df_all_row, user):
    """
    Generates and writes data files formatted for NEMS web app including study, heading,
    row, longthin, and footnote specifications.

    Each section of the code prepares data for a specific file:
    - db_study: Contains general study information.
    - db_heading: Contains headings for different tables in the study.
    - db_row: Contains detailed row information for tables.
    - db_longthin: Contains the 'long thin' formatted data for specific database operations.
    - db_footnote: Contains footnotes associated with the data tables.

    Parameters
    ----------
    table_spec : dict
        A dictionary containing specifications for the study, such as StudyID, LongTitle, case_descriptions,
        ScenLabel, Scenario, DateKey, and other parameters required to generate the files.
    df_all_row : DataFrame
        A pandas DataFrame containing all table row data that will be used to create the longthin database files.

    Returns
    -------
    None
        This function does not return any value. It writes the output directly to files.

    Note
    ----
        "Bonus rows" are never written to the "longthin" file.
    """

    # Output file name convention like below:
    # ref[yyyy].[mm][dd]a.db_footnote.txt
    # ref[yyyy].[mm][dd]a.db_heading.txt
    # ref[yyyy].[mm][dd]a.db_longthin.txt
    # ref[yyyy].[mm][dd]a.db_row.txt
    # ref[yyyy].[mm][dd]a.db_study.txt

    file_name = table_spec['ScenDate']

    # Create db_study file, e.g.,--------------------------------------
    #   StudyID|LongTitle|case_descriptions|ScenLabel|Scenario|DateKey|ScenDate
    #   AEO2023|Annual Energy Outlook 2023|Reference Case|Reference|ref2023|d020623a|ref2023.d020623a
    head_str = "StudyID|LongTitle|case_descriptions|ScenLabel|Scenario|DateKey|ScenDate"

    row_str = (
        table_spec["StudyID"]  # 'AEO2023'
        + "|"
        + table_spec["LongTitle"]  # 'Annual Energy Outlook 2023'
        + "|"
        + table_spec["case_descriptions"]  # 'Reference Case'
        + "|"
        + table_spec["ScenLabel"]  # 'Reference'
        + "|"
        + table_spec["Scenario"]  # 'ref2024'
        + "|"
        + table_spec["DateKey"]  # 'd101323b'
        + "|"
        + table_spec["ScenDate"]  # 'ref2023.d020623a'
    )

    # File name e.g., 'ref2023.0206a.db_study.txt'
    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name + ".study.txt")
    else:
        file_path = os.path.join(user.outputdir, file_name + ".study.txt")
    with open(file_path, "w") as file:
        file.write(head_str + "\n")
        file.write(row_str)

    # Create db_heading, e.g.,------------------------------------------------
    #   StudyID|TableNumber|SequenceNumber|Title|SubTitle|StubHead|RegionType
    #   AEO2023|001|  1|Table 1.  Total Energy Supply, Disposition, and Price Summary|(quadrillion Btu, unless otherwise noted)| Supply, Disposition, and Prices|0
    #   AEO2023|002|  2|Table 2.  Energy Consumption by Sector and Source|(quadrillion Btu, unless otherwise noted)| Sector and Source|1
    #   AEO2023|003|  3|Table 3.  Energy Prices by Sector and Source|(2022 dollars per million Btu, unless otherwise noted)| Sector and Source|1

    # # Get table format
    # # HD_num = 2 for longthin, 3 for RAN
    # df_format = normalize_format(table_spec['df_format'], HD_num=2)

    # Get HD and RH rows
    df_format = table_spec["df_format"]
    df_fmt = df_format[df_format["CM"].isin(["HD", "RH"])].copy()

    df_fmt = df_fmt.groupby(["TableNumber"])["Label"].apply("|".join).reset_index()
    df_fmt["TableNum"] = "TN " + df_fmt["TableNumber"].astype(str).str.zfill(3)

    # Get table input
    df_tab = table_spec["table_input_df"]

    # Get regionality
    df = pd.merge(
        df_fmt, df_tab, left_on="TableNum", right_on="TableID", how="left"
    ).fillna("")

    # TODO: check with EIA:
    regions = {"NATNAM": 0, "REGNAM": 1, "NERCNAM": 2}
    df["Reg"] = df["Region"].apply(
        lambda x: 0 if x == "NATNAM" else (1 if x == "REGNAM" else 2)
    )

    # Concat columns to get strings
    df = df.reset_index()

    # SequenceNumber is "1" based !
    df["index"] = df["index"] + 1

    df["Label"] = (
        table_spec["StudyID"]
        + "|"
        + df["TableNumber"].astype(str).str.zfill(3)
        + "|"
        + df["index"].astype(str).str.rjust(3)
        + "|"
        + "Table "
        + df["TableNumber"].astype(str)
        + ".  "
        + df["Label"]
        + "|"
        + df["Reg"].astype(str)
    )

    # Insert header row
    head_str = "StudyID|TableNumber|SequenceNumber|Title|SubTitle|StubHead|RegionType"
    df.loc[-1] = None
    df.index = df.index + 1
    df = df.sort_index()
    df.loc[0, "Label"] = head_str

    df = df[df["TableNumber"].isin([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,20,22,23,24,30,32,34,35,36,37,38,39,40,
                                                        41,42,43,45,46,47,48,49,50,51,52,53,54,55,56,57,58,60,62,67,69,70,71,72,73,76,77,78,
                                                        81,90,94,95,99,100,113,114,115,123,133,139,140,148])]
    
    # Save
    # file name e.g., 'ref2023.0206a.db_heading.txt'
    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name + ".heading.txt")
    else:
        file_path = os.path.join(user.outputdir, file_name + ".heading.txt")
    df["Label"].to_csv(
        file_path, header=False, index=False, quoting=csv.QUOTE_NONE, sep="\t"
    )

    # Create db_row, e.g.,------------------------------------------------
    # StudyID|TableNumber|RowNum|Stub|Font|RowFmt|GLabel|Gunits|DaType|SubDat|Sector|SubSec|Source|SubSrc|Geogr|VarName|Bonus|Irow
    # AEO2023|001|001|||2|||||||||||0|000
    # AEO2023|001|002|Production|Bold|2|||||||||||0|000
    # AEO2023|001|003|   Crude Oil and Lease Condensate|Plain|2|Total Energy : Production : Crude Oil and Lease Condensate|quads|supply|production|total energy||crude oil and lease condensate||united states|sup_prd_ten_NA_clc_NA_usa_qbtu|0|001

    # Get table format
    df = df_format[df_format["CM"].isin(["RL"])].copy()

    # DRForm character 6:
    #   Indicates Bonus Rows - Internal Only in published tables.
    #       0: regular publishable row
    #       1: line is omitted in formal tables and xml output files
    # df_fmt['Bonus'] = df_fmt['DRForm'].str[-1]
    df["Bonus"] = df["DRForm"].str[-1]

    # Remove 'Bonus' rows - "Bonus rows" are never written to the "longthin" file !!!
    df = df[df["Bonus"] == "0"]

    # # Some RL has blank DRForm
    # # Check if values in column 'DRForm' are blank and assign 'p15000' to blank values
    df["Font"] = df["DRForm"].apply(
        lambda x: (
            "Bold"
            if (isinstance(x, str) and x and x[0] == "B")
            else "Plain" if pd.isna(x) or not x else "Plain"
        )
    )

    df["Label"] = (
        table_spec["StudyID"]
        + "|"
        + df["TableNumber"].astype(str).str.zfill(3)
        + "|"
        + df["RowNum"].astype(int).astype(str).str.zfill(3)
        + "|"
        + df["Label"]
        + "|"
        + df["Font"].astype(str)
        + "|"
        + df["RowFmt"].astype(int).astype(str)
        + "|"
        + df["GLabel"]
        + "|"
        + df["Gunits"]
        + "|"
        + df["DaType"]
        + "|"
        + df["SubDat"]
        + "|"
        + df["Sector"]
        + "|"
        + df["SubSec"]
        + "|"
        + df["Source"]
        + "|"
        + df["SubSrc"]
        + "|"
        + df["Geography Override"]
        + "|"
        + df["VarName"]
        + "|"
        + df["Bonus"]
        + "|"
        + df["IROWS"].astype(str).str.zfill(3)
    )
    # Insert header row
    # import csv
    head_str = "StudyID|TableNumber|RowNum|Stub|Font|RowFmt|GLabel|Gunits|DaType|SubDat|Sector|SubSec|Source|SubSrc|Geogr|VarName|Bonus|Irow"
    df.loc[-1] = None
    df.index = df.index + 1
    df = df.sort_index()
    df.loc[0, "Label"] = head_str

    # Save
    # file name e.g., 'ref2023.0206a.db_row.txt'
    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name + ".row.txt")
    else:
        file_path = os.path.join(user.outputdir, file_name + ".row.txt")

    # Convert 'Label' column values to strings
    df["Label"] = df["Label"].astype(str)
    df = df[df["TableNumber"].isin([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,20,22,23,24,30,32,34,35,36,37,38,39,40,
                                                        41,42,43,45,46,47,48,49,50,51,52,53,54,55,56,57,58,60,62,67,69,70,71,72,73,76,77,78,
                                                        81,90,94,95,99,100,113,114,115,123,133,139,140,148])]
    # Save the 'Label' column values to a text file
    with open(file_path, "w") as file:
        file.write("\n".join(df["Label"].values))

    # Create db_longthin, e.g.,------------------------------------------------
    # TableNumber|RowNum|RegionNum|Year|Data|StudyID=AEO2023| Scenario=ref2023| Datekey=d020623a
    # 001|003| 0|2022|    24.586992
    # 001|003| 0|2023|    25.571161
    # 001|003| 0|2024|    26.285040

    # df = df_all_row[['TableNumber', 'RowNum','RegionNum'] + table_spec['years']].copy()
    df = df_all_row[
        ["TableNumber", "RowNum", "RegionNum"] + table_spec["print_years"]
    ].copy()
    df_pivot = df.melt(
        id_vars=["TableNumber", "RowNum", "RegionNum"],
        var_name="Year",
        value_name="Data",
    )
    df_sorted = df_pivot.sort_values(
        by=["TableNumber", "RowNum", "RegionNum"], ascending=True
    )

    head_str = (
        "TableNumber|RowNum|RegionNum|Year|Data|StudyID="
        + table_spec["StudyID"]
        + "| Scenario="
        + table_spec["Scenario"]
        + "| Datekey="
        + table_spec["DateKey"]
    )

    df_sorted[head_str] = (
        df_sorted["TableNumber"].astype(int).astype(str).str.zfill(3)
        + "|"
        + df_sorted["RowNum"].astype(int).astype(str).str.zfill(3)
        + "|"
        + df_sorted["RegionNum"].astype(int).astype(str).str.rjust(2)
        + "|"
        + df_sorted["Year"].astype(str)
        + "|"
        # + df_sorted['Data'].astype(str)
        # + df_sorted['Data'].round(6).map('{:.6f}'.format).astype(str).str.rjust(13)
        # + df_sorted['Data'].astype(str)#.str.rjust(13)
        + df_sorted["Data"].map("{:.6f}".format).str.rjust(13)
    )


    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name + ".longthin.txt")
    else:
        file_path = os.path.join(user.outputdir, file_name + ".longthin.txt")
    
    df_sorted = df_sorted[df_sorted["TableNumber"].isin([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,20,22,23,24,30,32,34,35,36,37,38,39,40,
                                                        41,42,43,45,46,47,48,49,50,51,52,53,54,55,56,57,58,60,62,67,69,70,71,72,73,76,77,78,
                                                        81,90,94,95,99,100,113,114,115,123,133,139,140,148])]
    
    print("Growth Rate Calculation for LongThin")

    new_rows = []  # List to store new rows to append to dataframe
    for table_num in df['TableNumber'].unique():
        table_df = df_sorted[df_sorted['TableNumber'] == table_num]

        for row_num in table_df['RowNum'].unique():
            row_df = table_df[table_df['RowNum'] == row_num]

            try:
                for reg_num in row_df["RegionNum"].unique():
                    reg_df = row_df[row_df["RegionNum"] == reg_num]
            
                    year_firstyear_data = reg_df[reg_df['Year'] == table_spec["growth_rate_start"]]['Data'].values
                    year_lastyear_data = reg_df[reg_df['Year'] == user.last_year]['Data'].values
            
                    if year_firstyear_data != 0 and year_lastyear_data.size != 0:
                        avg_annual_change = ((year_lastyear_data[0]  / year_firstyear_data[0]) ** (1 / (len(table_spec["print_years"]) - 1)) - 1) *100
                    else:
                        avg_annual_change = np.nan
                    new_row = {'TableNumber': table_num, 'RowNum': row_num, 'RegionNum': reg_num, 'Year': int(9999), 'Data': avg_annual_change}
                    new_rows.append(new_row)
            except:       
                year_firstyear_data = row_df[row_df['Year'] == table_spec["growth_rate_start"]]['Data'].values
                year_lastyear_data = row_df[row_df['Year'] == user.last_year]['Data'].values
            
                if year_firstyear_data != 0 and year_lastyear_data != 0:
                    avg_annual_change = ((year_lastyear_data[0]  / year_firstyear_data[0]) ** (1 / (len(table_spec["print_years"]) - 1)) - 1) *100
                else:
                    avg_annual_change = np.nan
                reg_num = 0
                new_row = {'TableNumber': table_num, 'RowNum': row_num, 'RegionNum': reg_num, 'Year': int(9999), 'Data': avg_annual_change}
                new_rows.append(new_row)

      
    new_rows = pd.DataFrame(new_rows)
    new_rows[head_str] = (
        new_rows["TableNumber"].astype(int).astype(str).str.zfill(3) +
        "|" +
        new_rows["RowNum"].astype(int).astype(str).str.zfill(3) +
        "|" +
        new_rows["RegionNum"].astype(int).astype(str).str.rjust(2) +
        "|" +
        new_rows["Year"].astype(str) +
        "|" +
        new_rows["Data"].map("{:.6f}".format).str.rjust(13)
        #new_rows["Data"].astype(str).str.rjust(13)
        )
    df_sorted = pd.concat([df_sorted, new_rows], ignore_index=True)
    df_sorted[head_str].replace(to_replace=[np.nan, np.inf, -np.inf, "          nan"], value="- -",regex=True, inplace=True)
    df_sorted = df_sorted.sort_values(
            by=["TableNumber", "RowNum", "RegionNum"], ascending=True
        )

    # Published Table 55 (layin 123) should not have regions in longthin. Removing them here.
    df_sorted = df_sorted[~((df_sorted["TableNumber"] == 123) & (df_sorted["RegionNum"] > 0))]


    df_sorted[head_str].to_csv(
        file_path,
        header=True,
        index=False,
        quoting=csv.QUOTE_NONE,
        sep="\t",
        float_format="%.6f",
    )

    # Create db_footnote, e.g.,------------------------------------------------
    # StudyID|TableNumber|Footnum|Footnote
    # AEO2023|001|001|1/ Includes waste coal.
    # AEO2023|001|002|2/ These values represent the energy obtained from uranium when it is used in light water reactors.  The total energy content of uranium

    df = df_format[df_format["CM"].isin(["FN"])].copy()

    # Count Footnum by table
    df["Footnum"] = df.groupby("TableNumber").cumcount() + 1

    # Concat columns to get strings
    df = df.reset_index()
    df["Label"] = (
        table_spec["StudyID"]
        + "|"
        + df["TableNumber"].astype(str).str.zfill(3)
        + "|"
        + df["Footnum"].astype(str).str.zfill(3)
        + "|"
        + df["Label"]
    )

    # Insert header row
    head_str = "StudyID|TableNumber|Footnum|Footnote"
    df.loc[-1] = None
    df.index = df.index + 1
    df = df.sort_index()
    df.loc[0, "Label"] = head_str

    for line in df:
        columns = line.split('|')


    # Save file, name e.g., 'ref2023.0206a.db_heading.txt'
    if user.outputdir == "output":
        file_path = os.path.join(os.getcwd(), "output", file_name + ".footnote.txt")
    else:
        file_path = os.path.join(user.outputdir, file_name + ".footnote.txt")

    df = df[df["TableNumber"].isin([1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,20,22,23,24,30,32,34,35,36,37,38,39,40,
                                                        41,42,43,45,46,47,48,49,50,51,52,53,54,55,56,57,58,60,62,67,69,70,71,72,73,76,77,78,
                                                        81,90,94,95,99,100,113,114,115,123,133,139,140,148])]
    
    df["Label"].to_csv(
        file_path, header=False, index=False, quoting=csv.QUOTE_NONE, sep="\t"
    )

    
    return


def read_tabreq(path):
    """
    Reads the tabreq.txt to identify and extract table numbers that are marked as required.
    The function skips the initial part of the file which includes headers or metadata, and
    processes each line to determine whether the table described in that line is required.

    tabreq.txt (from line 7):
        ----------------------------------------------------------
        1   1. Total Energy Supply, Disposition, and Price Summary
        1   2. Energy Consumption by Sector and Source
        1   3. Energy Prices by Sector and Source
               ...
        0  27. Non-Utility Electricity Capacity
               ...
        1 150. Convergence Indicators
        ----------------------------------------------------------
        1: required
        0: not required

    If a table is required, its table number is extracted, formatted, and added to a list.

    Parameters
    ----------
    path : str
        The path to the tabreq.txt

    Returns
    -------
    list of str
        A list of table id of all required tables. e.g., ['TN 001', 'TN 002', ...]
    """

    # Open the file in read mode
    with open(path, "r") as file:
        # Skip the first 6 lines
        for _ in range(6):
            next(file)

        # Read the remaining lines and process each line
        table_ids_req = []
        for line in file:
            line = line.strip()
            # Extract the desired parts from each line
            required = line[0]
            table_num = line[2:5].strip()  # Remove leading whitespace
            # table_name = line[6:]
            # Append the extracted parts to the list if required is 1
            if required == "1":
                table_ids_req.append(f"TN {table_num.strip().zfill(3)}")

    return table_ids_req


@log_execution(main_logger, message="report_main")
def main(user):
    """
    This is the main body of the NEMS RW Prototype, performing several tasks
    and returning three objects: d_base, d_formatted, and df_all_row.

    Tasks performed:
    1. Create table specifications (metadata) - 'init'
    2. Perform SME calculations - 'preprocessor_sme'
    3. Create base tables - 'make_base_tables'
    4. Perform additional calculations - 'postprocessor_base'
    5. Create formatted tables - 'make_formatted_tables'
    6. Create all_row table & write to Excel file - 'make_all_row_table'
    7. Create published AEO tables & write to Excel files - 'create_aeo_files'
    8. Create and write longthin txt files - 'write_longthin'
    9. Generate RAN file by calling - 'make_ran'

    Returns:
    d_base : dict
        A dictionary containing the base tables.
    d_formatted : dict
        A dictionary containing the formatted tables.
    df_all_row : pd.DataFrame
        A table containing all rows from all formatted tables.
    """

    main_logger.info("\n============ main ============")

    table_spec = {}
    d_base = {}
    d_formatted = {}
    df_all_row = pd.DataFrame()

    # 1. Init to create table spec (metadata)
    main_logger.info("\n============ init ============")
    table_spec, dat = init(user)

    # 2. Run preprocessor_sme
    main_logger.info("\n============ preprocessor_sme ============")

    # Create a wrapper around the preprocessor_sme function, and apply
    # the log decorator to log the execution of the decorated function
    @log_execution(main_logger, message="preprocessor_sme")
    def decorated_preprocessor_sme(table_spec):
        return preprocessor_sme(table_spec)

    dfd_sme = decorated_preprocessor_sme(table_spec)

    # 3. Create base tables
    main_logger.info("\n============ make_base_tables ============")
    d_base = make_base_tables(table_spec, dfd_sme, user)

    # 4. Post fix base tables
    main_logger.info("\n============ postprocessor_base ============")

    # Create a wrapper around the postprocessor_base function, and apply
    # the log decorator to log the execution of the decorated function
    @log_execution(main_logger, message="postprocessor_base")
    def decorated_postprocessor_base(table_spec, d_base, dfd):
        return postprocessor_base(table_spec, d_base, dfd)

    d_base_fixed = decorated_postprocessor_base(table_spec, d_base, dfd)

    # 5. Create formatted tables
    main_logger.info("\n============ make_formatted_tables ============")
    d_formatted = make_formatted_tables(table_spec, d_base_fixed)

    # 6. Create the all_row_csv table and write to csv file
    main_logger.info("\n============ make_all_row_table ============")
    df_all_row = make_all_row_table(table_spec, d_formatted,user)

    # 7. Create all published AEO tables and write to Excel files
    main_logger.info("\n============ create_aeo_files ============")
    create_aeo_files(table_spec, d_formatted, user)  # , df_all_row)

    # 8. Create and write longthin txt files
    if int(table_spec["settings"]["create_longthin_files"]) == 1:
        main_logger.info("\n============ write_longthin ============")
        write_longthin(table_spec, df_all_row, user)

    # 9. Generate the RAN file
    if int(table_spec["settings"]["create_ran_file"]):
        main_logger.info("\n============ make_ran ============")
        main_logger.info("\n------------ (see RAN_log) ------------")
        make_ran(
            table_spec["all_row_file_path"],
            table_spec["df_format"],
            table_spec["RSVer"],
            table_spec["RSFile"],
            table_spec["RSScen"],
            table_spec["ScenDate"],
            table_spec["first_year"],
            table_spec["last_year"],
            user,
        )

    # main_logger.info('\n============ all done! ============')
    
    # 10. Respecify Columns for api.csv file to use with Validator
    # open table_spec["all_row_file_path"]
    df = pd.read_csv(table_spec["all_row_file_path"])
    
    # change columns (insert columns here) to correct format and name
    df['TableNumber'] = df['TableNumber'].astype(int)
    df['RowNum'] = df['RowNum'].astype(int)
    df['RegionNum'] = df['RegionNum'].astype(int)

    df['TableNumber'] = df['TableNumber'].astype(str)
    df['RowNum'] = df['RowNum'].astype(str)
    df['RegionNum'] = df['RegionNum'].astype(str)

    # read in layin.csv and grab RN rows
    dflayin = pd.read_csv("input/layin.csv")
    dflayin_tabname = dflayin[dflayin["CM"] == "RN"]
    dflayin_tabname['Table Number'] = dflayin_tabname['Table Number'].astype(int)
    dflayin_tabname['Table Number'] = dflayin_tabname['Table Number'].astype(str)
    dflayin_tabname = dflayin_tabname[["CM", "Row Label", "Table Number"]]

    for idx,row in df.iterrows():
        match = dflayin_tabname[dflayin_tabname['Table Number'] == row['TableNumber']]

        if not match.empty:
            row_label = match.iloc[0]['Row Label']
            df.at[idx, 'VarNam2'] = f"{row_label}000:{row['VarNam2']}"

    # save over file
    df.to_csv(table_spec["all_row_file_path"], index=False)

    return d_base, d_formatted, df_all_row


def get_vartable(arr):
    """Get _variable listing and convert to pd.DataFrame

    Parameters
    ----------
    arr: np.ndarray
        _variablelisting from restart file
    """
    columns = [
        "Fortran Variable Name",
        "Common Block Name",
        "Dimension Params",
        "Extra Indices",
    ]
    vartable = pd.DataFrame(arr, columns=columns)
    vartable["Fortran Variable Name"] = vartable["Fortran Variable Name"].str.upper()
    vartable = vartable.set_index("Fortran Variable Name")
    vartable["Fortran Variable Name"] = vartable.index
    return vartable


def to_pandas(key, arr, user):
    """
    Takes in a numpy array of any dimension and outputs a DataFrame or Series

    Parameters
    ----------
    key: str
        Common Block/Variable Name
    arr: np.ndarray
        Array from restart file

    Returns
    -------
    A pandas DataFrame with of the numpy array
    """
    vartable = user.file_map["vartable"]
    comblock = key.split("/")[0]  # Get common block
    var = key.split("/")[1].upper()  # Get variable name
    vardims = vartable.loc[var]

    # Get first and last year to trim columns if needed
    fyr, lyr = user.first_year, user.last_year
    yearcols = range(fyr, lyr + 1)

    # Quick check for if vardim holds the variable twice. If so, select the correct one from the common block we want
    if isinstance(vardims, pd.DataFrame):
        vardims = vardims.loc[vardims["Common Block Name"] == comblock]
    else:
        # Making vardims a pandas dataframe for easier search and work with.
        vardims = pd.DataFrame(vardims)
        vardims = vardims.transpose(copy=True)
    dparams = vardims.loc[var]["Dimension Params"]

    if isinstance(dparams, str) and dparams[0] == "[":
        dparams = dparams.strip("[']").split("', '")

    indices = [[_ for _ in range(1, int(s) + 1)] for s in arr.shape]
    dims = arr.ndim

    # Year variables with start year
    yearvars = ["MNUMYR", "MNUMY5", "MNMYRF", "MECPYR", "MNXYRS"]

    # Process column names
    buffer = []  # Holds list of columns
    for i, param in enumerate(dparams):
        if param.isdigit():  # Prefix 'M' to integer column names
            dparams[i] = f"M{param}"

        # If duplicate columns, rename to avoid errors
        if param in buffer:
            dparams[i] = f"{dparams[i]}_{buffer.count(param)}"
        buffer.append(param)  # List of variable names used

        # Change years to 1990=index instead of 1-indexed
        if param in yearvars:
            start = user.first_year - 1  # 1-indexed
            indices[i] = [index + start for index in indices[i]]
    if not indices:  # For blank indicies
        indices = [[1]]  # Create 1-value list for 0D tables

    # 0D and 1D arrays convert to Series directly
    if dims < 2:
        ser = pd.Series(arr, indices[0], name=key, dtype=arr.dtype).rename_axis(
            dparams[0]
        )
        if dparams[0] in yearvars:
            ser = ser.loc[fyr:lyr]
        return ser

    # 2D arrays convert to DataFrames directly
    elif dims == 2:
        df = pd.DataFrame(arr, index=indices[0], columns=indices[1])
        for i, param in enumerate(dparams):
            df = df.rename_axis(param, axis=i)

        # Make sure Years are always columns; trim to first/last year
        df = df.T if dparams[0] in yearvars else df
        df = df[yearcols] if set(yearvars).intersection(set(dparams)) else df

    # 3D or greater arrays need to be processed.
    else:
        # Create index using cartesian product of list of ranges
        index = pd.MultiIndex.from_product(indices, names=dparams)

        flat = np.ravel(arr, order="C").copy()

        # Create DataFrame from flattened array with MultiIndex
        df = pd.DataFrame(flat, index=index, columns=[""])

        is_years = set(yearvars).intersection(set(dparams))
        if is_years:  # Pivot if data is by year
            year = is_years.pop()
            dparams.remove(year)
            df = df.reset_index().pivot(values="", columns=year, index=dparams)[
                yearcols
            ]
    return df.sort_index()


def read_npz(user):
    """Reads the data from the restart file restart.npz into memory.

    Parameters
    ----------
    user : user
        User object containing user-specific information - see user class

    Returns
    -------
    dataframe
        Dictionary of numpy arrays converted to DataFrames
    """
    npz = user.file_map["RestartNPZ"]
    vartable = user.file_map["vartable"]
    # Load npz file into dictionary of numpy arrays
    npd = {}
    with np.load(npz, allow_pickle=True) as f:
        for key, val in f.items():
            npd[key] = val

    # Get variable listing from restart file -> DataFrame
    arr = npd.pop("_variablelisting")  # Removal avoids error in key.split() ↓
    vartable = get_vartable(arr)

    errors = {}  # Holds variables with errors
    dfd = {}  # DataFrame dictionary

    for key, arr in npd.items():
        try:
            key = key.upper()
            df = to_pandas(key, arr, user)
            dfd[key] = df

            # Duplicate keys to variable names w/o common block for easier access
            abbrev = key.split("/")[1].upper()
            if dfd.get(abbrev) is not None:
                df = df if abbrev.startswith("A") else dfd[abbrev]
            dfd[abbrev] = df

            # Remove key if previously had an error but now succesful
            if key in errors:
                errors.pop(key)
        except Exception as e:
            # VarIssue holds any variables that may have an issue
            print(f"{key} error: {e}")
            errors[key] = traceback.format_exception_only(e)

    # Output errors to a csv file
    err_df = pd.DataFrame([key, val] for key, val in errors.items())
    if not err_df.empty:
        err_df.to_csv("npz_errors.csv")

    return dfd


def load_NEMSVardf_from_file(locofdict):
    f = open(locofdict, "r")
    data = f.read()
    f.close()
    return eval(data)


def run(SCEDES, curirun):
    """_summary_

    Note: cwd must be the Reporter directory.

    Parameters
    ----------
    SCEDES : _type_
        _description_
    """
    sys.stdout.flush()
    sys.stderr.flush()

    # pass
    # print('Start running NEMS reporter. It may take five minutes ...')
    # print('For execution details, please see RW_log')
    # print('Hello from RW_reporter_main.py!')
    # print(os.getcwd())

    class user:
        pass

    user.file_map = {}

    # Case descriptions used by longthin
    # TODO: check with EIA to get "accurate" case descriptions
    # {ScenLabel, case_descriptions}
    
    # Set Scenario and Datekey from scedes file OR from SCEDES line in main if run in standalone
    user.Scenario = SCEDES["SCEN"]
    user.DateKey = SCEDES["DATE"]
    user.ScenDate = SCEDES["SCEN"] + "." + SCEDES["DATE"]  # ref2024.d030424d'

    # Projection years
    user.first_year = int(SCEDES["FIRSYR"])
    user.last_year = int(SCEDES["LASTYR"])

    # Setting all inputs via config.ini
    inputpath = {}
    for key, value in config.items("inputpath"):
        inputpath[key] = value

    outputpath = {}
    for key, value in config.items("outputpath"):
        outputpath[key] = value

    longthin = {}
    for key, value in config.items("longthin"):
        longthin[key] = value

    casedescription = {}
    for key, value in config.items("casedescriptions"):
        casedescription[key] = value

    scenlabel = {}
    for key, value in config.items("scenlabel"):
        scenlabel[key] = value


    # Input files
    user.restartnpz_path = inputpath["restartnpz"]
    user.table_var_def_path = inputpath["tablevardeftoml"]
    user.layin_ver = "csv"  # switch: csv / xls
    user.format_path = inputpath["layin"]
    user.format_path_csv = inputpath["layin"]
    user.table_input_path = inputpath["table_input"]
    user.region_input_path = inputpath["regional_input"]
    user.par_path = inputpath["parameter_path"]
    user.coefficients_path = inputpath["coefficients_path"]
    user.citation_path = inputpath["citation_path"]
    user.tabreq_path = inputpath["tabreq_path"]
    user.table_mapping_path = inputpath["table_mapping"]
    user.nemsvardf_path = inputpath["nemsvardf"]
    user.outputdir = outputpath["outputpath"]
    
    BASE_DIR = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    if user.restartnpz_path == "restart.npz":
        user.file_map["RestartNPZ"] = BASE_DIR + "\\" + user.restartnpz_path
    else:
        print(user.file_map["RestartNPZ"])
    if user.nemsvardf_path == "NEMSVardf.csv":
        user.file_map["NEMSVardf"] = BASE_DIR + "\\" + user.nemsvardf_path
    else:
        user.file_map["NEMSVardf"] = user.nemsvardf_path
    user.cyclenum = int(curirun)
    vartable = pd.read_csv(user.file_map["NEMSVardf"])
    vartable.set_index(vartable["Fortran Variable Name"], inplace=True)
    user.file_map["vartable"] = vartable

    # NEMS info
    # For longthin db_study e.g.,
    #   StudyID|LongTitle|case_descriptions|ScenLabel|Scenario|DateKey|ScenDate
    #   AEO2023|Annual Energy Outlook 2023|Reference Case|Reference|ref2023|d020623a|ref2023.d020623a
    # TODO: need to think about the StudyID, LongTitle, ScenLabel, and case_descriptions
    user.StudyID = longthin["study_id"] + longthin["year"]
    user.LongTitle = "Annual Energy Outlook " + longthin["year"]
    try:
        user.ScenLabel = scenlabel[user.Scenario]
    except KeyError:
        user.ScenLabel = user.Scenario
    # user.case_descriptions = case_descriptions[user.ScenLabel] # 'Reference Case' ?
    user.case_descriptions = user.ScenLabel
    user.release_date = scenlabel["release_date"]

    # For RAN
    user.RSVer = "AA01"  # Scenario Header data
    user.RSFile = user.file_map["RestartNPZ"]
    user.RSScen = "../../" + user.Scenario + "/" + user.DateKey + "/" + "RESTART.unf"

    # Set tables
    blank_tables = [63, 64, 82, 84, 85, 86, 87, 88, 89, 92, 104, 109, 120, 121]
    under_development_tables = [31, 125]  # will be "zero out" onec all done

    user.unavailable_tables = blank_tables + under_development_tables

    NEMS_d_base, NEMS_d_formatted, NEMS_df_all_row = main(user)

    # TODO: return something meaningful


if __name__ == "__main__":
    # Set these values if running standalone, else files will return test.d000000a
    SCEDES = {"SCEN": "test", "DATE": "d000000a", "FIRSYR": 1990, "LASTYR": 2050}
    run(SCEDES, 1)
    print("Done!")